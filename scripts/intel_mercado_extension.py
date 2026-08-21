# -*- coding: utf-8 -*-
"""
Extensão do extract_data.py: Inteligência de Mercado (visitação de concorrentes por região)
=============================================================================================

CONTEXTO
--------
Pedido do usuário (21/08/2026): nova aba "Inteligência de Mercado" no Performance Dashboard,
com a visitação mensal dos nossos parques comparada com a dos concorrentes diretos de cada
região -- mesma comparação que ele hoje monta manualmente em PPT pra apresentações.

Regiões e "réguas" (unidade de comparação -- normalmente 1 por região, com TODOS os parques
do grupo daquela região juntos + todos os concorrentes da região; Rio de Janeiro é a EXCEÇÃO
pedida explicitamente pelo usuário: AquaRio e Paineiras, mesmo sendo os dois no Rio, têm
concorrentes diferentes entre si e por isso viram DUAS réguas separadas, não uma só):

  - Rio de Janeiro:
      régua "AquaRio": AquaRio (nosso) x YUP Star, Museu do Amanhã, BioParque
      régua "Paineiras": Paineiras (nosso) x Trem do Corcovado, Bondinho Pão de Açúcar
  - Aparecida (SP):
      régua "Três Pescadores": Três Pescadores (nosso) x Trem do Devoto, Santuário Nacional
      Aparecida, Cidade do Romeiro
  - Curitiba:
      régua "Vila Velha": Vila Velha (nosso) x Trem de Curitiba, Buraco do Padre, Het Dorp,
      Museu Oscar Niemeyer (MON)
  - Foz do Iguaçu:
      régua "Foz do Iguaçu": AquaFoz, Marco das 3 Fronteiras e PNI (nossos, os 3 juntos) x
      Parque das Aves, Turismo Itaipu, Wonder Park Foz, Dreams Park Show

FONTES: 4 planilhas Google Sheets fixas (uma por região), compartilhadas com a service account
do pipeline -- ids em cfg["intel_mercado_ids"]["rio"/"aparecida"/"curitiba"/"foz"]. Os dados só
chegam fechados até o último mês (o usuário recebe após o fechamento de cada mês) -- é normal
ter meses futuros/vigente vazios (fica None, não 0 -- ver _to_num).

NÚMEROS DOS NOSSOS PARQUES: NÃO são lidos destas planilhas (mesmo quando a planilha traz uma
linha com o nosso próprio parque, ex. "AquaRio", "Vila Velha", "AQUAFOZ" -- essas linhas
existem só pra composição visual/cálculo de Share dentro da própria planilha do usuário, e
podem ficar defasadas em relação à fonte oficial, ex. antes de um ajuste retroativo). O painel
usa o número oficial já existente em data["VISITACAO"][MES]["summary"][parque]["realizado"]
pra cada um dos nossos parques -- evita ter duas fontes de verdade pro mesmo dado. Esta
extensão só devolve os números dos CONCORRENTES; build_intel_mercado() junta os dois na hora
de montar a saída final.

REGRA BIOPARQUE: BioParque aparece como concorrente na régua AquaRio, mas segue a MESMA regra
já usada no resto do painel (ver BIOPARQUE_SAIDA_ANO_MES / _bioparque_ainda_conta no
extract_data.py) -- a partir de Agosto/2026 o valor mensal vem sempre None aqui, INDEPENDENTE
do botão "Ocultar Bio" (que é um filtro só do front-end, aplicado em cima disso). Ou seja,
mesmo que o usuário reative o botão (voltar a mostrar Bio), esta aba nunca mostra Bio além de
Julho/2026 -- pedido explícito do usuário: "mesmo com botão só considerar até Julho quando
inativo".

ESTRUTURA DAS PLANILHAS RIO / APARECIDA / CURITIBA (aba "2025x2026", idêntica nas 3): blocos
de 3 linhas por entidade (1a linha = nome na coluna B + ano 2025 na coluna C + valores 2025
nas colunas D..O; 2a linha = ano 2026 na coluna C + valores 2026 (os que usamos) nas colunas
D..O; 3a linha = "%" -- ignorada, recalculada no front-end a partir dos números crus). Cada
planilha tem blocos extras depois (histórico de anos anteriores, Share já calculado) que são
ignorados de propósito -- a leitura para no primeiro espaço em branco após o primeiro bloco
encontrado. Ver _ler_blocos_ano_a_ano().

ESTRUTURA DA PLANILHA FOZ (aba "2024x2025X2026"): 7 blocos sequenciais de ~25 linhas, um por
entidade (AQUAFOZ, M3F, PARQUES DAS AVES, TURISMO ITAIPU, PARQUE NACIONAL DO IGUAÇU, Wonder
Park Foz, Dreams Park Show), cada um com 5 categorias de origem (Morador/Estado/Brasileiro/
Mercosul/Estrangeiro) x (2024/2025/2026/%) e, no fim do bloco, uma linha "TOTAL 2026" (e às
vezes "TOTAL 2025") já somada. Por pedido do usuário ("pode somar o total, só compare se está
batendo com o número total de visitação do parque" -- confirmado manualmente: bate com o
número oficial de visitação do parque nos meses fechados), usa-se essa linha TOTAL já pronta
em vez de somar as 5 categorias de novo. Só os blocos de CONCORRENTE são lidos daqui (Parque
das Aves, Turismo Itaipu, Wonder Park Foz, Dreams Park Show) -- AQUAFOZ/M3F são ignorados por
serem nossos (ver acima), e "Parque Nacional do Iguaçu" (nosso -- PNI) está com o bloco
zerado nesta planilha, então também é ignorado aqui (o número do PNI vem da fonte oficial).

Turismo Itaipu só tem dado até Março/2026 nesta planilha (confirmado com o usuário que é o que
a fonte tem -- meses seguintes ficam None, não 0, pra não sugerir "visitação zero"). Wonder
Park Foz e Dreams Park Show estão totalmente vazios nesta planilha até a data desta
implementação (21/08/2026) -- o usuário inclusive não conhecia o "Dreams Park Show" como
concorrente até ver esta planilha; ficam com todos os meses None e aparecem prontos pro filtro
de região assim que a fonte trouxer dado.

NENHUMA planilha aqui pode derrubar o pipeline inteiro se der erro -- build_intel_mercado() é
sempre chamado de dentro de um try/except em main(), com um fallback seguro (réguas com
mensal vazio) e "AVISO: ..." no stderr, igual às outras extensões (cross_aquafoz_extension.py,
top5_origem_extension.py). Ver "COMO INTEGRAR" no fim do arquivo.
"""

import unicodedata

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def _norm(s):
    """Normaliza rótulo de célula (nome de linha) pra comparação: remove acento, baixa a
    caixa, colapsa espaço. Não usado para os VALORES numéricos -- ver _to_num."""
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.lower().split())


def _to_num(v):
    """Converte valor de célula pra número, tratando 'NA'/'N/A'/'#DIV/0!'/'-'/vazio como None
    (em vez de virar 0 escondido) -- importante pro front-end saber quando o dado simplesmente
    não existe ainda (ex.: mês futuro, ou concorrente sem dado nesse mês)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s or s in ("-",) or s.upper() in ("NA", "N/A", "#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return None
    try:
        return float(s)
    except ValueError:
        try:
            return float(s.replace(".", "").replace(",", "."))
        except ValueError:
            return None


def _download_workbook(drive_service, spreadsheet_id):
    """Baixa a planilha (Drive API) e abre com openpyxl -- mesmo padrão de
    cross_aquafoz_extension.py / top5_origem_extension.py (self-contained, não depende do
    cache/import de extract_data.py, pra evitar import circular)."""
    import io
    import openpyxl
    from googleapiclient.http import MediaIoBaseDownload

    meta = drive_service.files().get(
        fileId=spreadsheet_id, fields="mimeType", supportsAllDrives=True
    ).execute()
    is_native_sheet = meta["mimeType"] == "application/vnd.google-apps.spreadsheet"

    buf = io.BytesIO()
    if is_native_sheet:
        request = drive_service.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive_service.files().get_media(fileId=spreadsheet_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(buf, request, chunksize=10 * 1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=5)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True, read_only=True)


def _get_rows(wb, sheet_name):
    ws = wb[sheet_name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


# ---------------------------------------------------------------------------------------------
# Rio / Aparecida / Curitiba -- aba "2025x2026", blocos de 3 linhas (nome+2025, 2026, %)
# ---------------------------------------------------------------------------------------------

def _ler_blocos_ano_a_ano(rows, rotulo_para_entidade):
    """Lê os blocos de 3 linhas das planilhas Rio/Aparecida/Curitiba (ver docstring do
    módulo). Devolve {entidade: [12 valores mensais, Jan..Dez, None se vazio]}, só para os
    rótulos presentes em rotulo_para_entidade (outros blocos -- nossos parques, histórico,
    Share já calculado -- são ignorados de propósito). Para de procurar assim que encontrar
    uma linha em branco depois do primeiro bloco reconhecido (fim da tabela principal, antes
    de qualquer bloco secundário de histórico/Share)."""
    out = {}
    i = 0
    n = len(rows)
    achou_alguma = False
    while i < n:
        row = rows[i]
        if not row or len(row) < 3:
            if achou_alguma:
                break
            i += 1
            continue
        rotulo, ano = row[1], row[2]
        if isinstance(rotulo, str) and ano == 2025.0 and i + 1 < n:
            entidade = rotulo_para_entidade.get(_norm(rotulo))
            row2026 = rows[i + 1]
            if entidade and row2026 and len(row2026) > 2 and row2026[2] == 2026.0:
                valores = [_to_num(row2026[3 + m]) if 3 + m < len(row2026) else None for m in range(12)]
                out[entidade] = valores
                achou_alguma = True
            i += 3
            continue
        if achou_alguma and rotulo is None and ano is None:
            break
        i += 1
    return out


RIO_ROTULOS = {
    "trem": "Trem do Corcovado",
    "bondinho": "Bondinho Pão de Açúcar",
    "yup star": "YUP Star",
    "museu do amanha": "Museu do Amanhã",
    "bioparque do rio": "BioParque",
    # "paineiras" e "aquarico" -- nossos, ignorados (ver docstring do módulo)
}

APARECIDA_ROTULOS = {
    "trem do devoto": "Trem do Devoto",
    "santuario": "Santuário Nacional Aparecida",
    "cidade do romeiro": "Cidade do Romeiro",
}

CURITIBA_ROTULOS = {
    "trem de curitiba": "Trem de Curitiba",
    "buraco do padre": "Buraco do Padre",
    "het dorp": "Het Dorp",
    "mon": "Museu Oscar Niemeyer",
}


# ---------------------------------------------------------------------------------------------
# Foz -- aba "2024x2025X2026", blocos de ~25 linhas com linha "TOTAL 2026" já somada
# ---------------------------------------------------------------------------------------------

FOZ_ROTULOS = {
    "parques das aves": "Parque das Aves",
    "turismo itaipu": "Turismo Itaipu",
    "wonder park foz": "Wonder Park Foz",
    "dreams park show": "Dreams Park Show",
    # AQUAFOZ / M3F / PARQUE NACIONAL DO IGUACU -- nossos, ignorados (ver docstring do módulo)
}


def _ler_blocos_foz(rows, rotulo_para_entidade):
    """Lê os blocos da aba Foz: acha a linha de cabeçalho de cada bloco (coluna A = nome da
    entidade) e, dentro do bloco (até o próximo cabeçalho), a linha 'TOTAL 2026' com os 12
    valores mensais já somados nas colunas C..N. Devolve {entidade: [12 valores mensais]} --
    bloco totalmente vazio/zerado (concorrente sem nenhum dado na planilha, ex. Wonder Park
    Foz/Dreams Park Show até a data desta implementação) devolve [None]*12, não [0]*12, pra
    não sugerir "visitação zero" no front-end."""
    out = {}
    n = len(rows)
    # Cabeçalho de bloco = coluna A com o nome da entidade E coluna B vazia (as linhas de
    # categoria dentro do bloco, ex. 'MORADOR'/'ESTADO (Paraná)'/..., têm coluna A também como
    # texto, mas SEMPRE com o ano 2024.0 na coluna B -- por isso o filtro de r[1] is None é
    # essencial pra não confundir categoria com cabeçalho de bloco).
    idx_cabecalhos = [
        i for i, r in enumerate(rows)
        if r and isinstance(r[0], str) and r[0].strip() and (len(r) < 2 or r[1] is None)
    ]
    for pos, i in enumerate(idx_cabecalhos):
        entidade = rotulo_para_entidade.get(_norm(rows[i][0]))
        if not entidade:
            continue
        fim = idx_cabecalhos[pos + 1] if pos + 1 < len(idx_cabecalhos) else n
        total_row = None
        for j in range(i, min(fim, n)):
            r = rows[j]
            if r and len(r) > 1 and isinstance(r[1], str) and _norm(r[1]) == "total 2026":
                total_row = r
                break
        if not total_row:
            continue
        # 0 aqui sempre significa "mês sem dado ainda" (soma de categorias todas vazias), não
        # "visitação zero" -- um parque/atrativo em operação nunca fecha o mês com 0 visitantes
        # -- por isso cada 0 é convertido pra None individualmente (não só quando o bloco
        # inteiro está zerado, ver Turismo Itaipu: Jan-Mar com dado real, Abr em diante 0).
        valores = [
            (None if (v is None or v == 0) else v)
            for v in (_to_num(total_row[2 + m]) if 2 + m < len(total_row) else None for m in range(12))
        ]
        out[entidade] = valores
    return out


# ---------------------------------------------------------------------------------------------

def _aplica_regra_bioparque(valores_mensais, bioparque_ainda_conta_fn):
    """Zera (None) os meses a partir de quando o BioParque 'sai' da regra (ver
    _bioparque_ainda_conta em extract_data.py), independente do botão do front-end -- ver
    docstring do módulo."""
    return [
        v if bioparque_ainda_conta_fn(2026, m) else None
        for m, v in enumerate(valores_mensais, start=1)
    ]


# Réguas fixas -- ver docstring do módulo pra explicação de cada uma, em especial a exceção
# do Rio (AquaRio e Paineiras viram 2 réguas, não 1, mesmo sendo a mesma região).
REGUAS = [
    {
        "id": "aquario", "regiao": "Rio de Janeiro", "nome": "AquaRio",
        "nossoParques": ["AquaRio"],
        "concorrentes": ["YUP Star", "Museu do Amanhã", "BioParque"],
    },
    {
        "id": "paineiras", "regiao": "Rio de Janeiro", "nome": "Paineiras",
        "nossoParques": ["Paineiras"],
        "concorrentes": ["Trem do Corcovado", "Bondinho Pão de Açúcar"],
    },
    {
        "id": "aparecida", "regiao": "Aparecida", "nome": "Três Pescadores",
        "nossoParques": ["Três Pescadores"],
        "concorrentes": ["Trem do Devoto", "Santuário Nacional Aparecida", "Cidade do Romeiro"],
    },
    {
        "id": "curitiba", "regiao": "Curitiba", "nome": "Vila Velha",
        "nossoParques": ["Vila Velha"],
        "concorrentes": ["Trem de Curitiba", "Buraco do Padre", "Het Dorp", "Museu Oscar Niemeyer"],
    },
    {
        "id": "foz", "regiao": "Foz do Iguaçu", "nome": "Foz do Iguaçu",
        "nossoParques": ["AquaFoz", "M3F", "PNI"],
        "concorrentes": ["Parque das Aves", "Turismo Itaipu", "Wonder Park Foz", "Dreams Park Show"],
    },
]


def build_intel_mercado(drive_service, ids, visitacao_por_mes, bioparque_ainda_conta_fn):
    """Função principal -- chamar de dentro de main() do extract_data.py e salvar o resultado
    em data['INTEL_MERCADO'].

    ids: dict com as 4 chaves de config.json (cfg["intel_mercado_ids"]["rio"/"aparecida"/
    "curitiba"/"foz"]).
    visitacao_por_mes: o próprio dict que já vai em data['VISITACAO'] (MES em caixa alta ->
    "summary" -> parque -> "realizado") -- usado pra pegar o número OFICIAL dos nossos
    parques (ver docstring do módulo -- não lemos o número dos nossos parques destas 4
    planilhas novas).
    bioparque_ainda_conta_fn: a função _bioparque_ainda_conta já existente em extract_data.py.

    Devolve {'reguas': [...]} -- cada régua com id/regiao/nome/nossoParques/concorrentes e
    'mensal': {mes: {entidade: valor ou None}} (mês em Title Case, ex. 'Janeiro') juntando
    nosso(s) parque(s) (fonte oficial) e concorrentes (das 4 planilhas novas). O Share/Captação
    (nosso ÷ concorrente) é calculado no front-end a partir destes números crus, não aqui.
    """
    wb_rio = _download_workbook(drive_service, ids["rio"])
    wb_aparecida = _download_workbook(drive_service, ids["aparecida"])
    wb_curitiba = _download_workbook(drive_service, ids["curitiba"])
    wb_foz = _download_workbook(drive_service, ids["foz"])

    concorrentes = {}
    concorrentes.update(_ler_blocos_ano_a_ano(_get_rows(wb_rio, "2025x2026"), RIO_ROTULOS))
    if "BioParque" in concorrentes:
        concorrentes["BioParque"] = _aplica_regra_bioparque(concorrentes["BioParque"], bioparque_ainda_conta_fn)
    concorrentes.update(_ler_blocos_ano_a_ano(_get_rows(wb_aparecida, "2025x2026"), APARECIDA_ROTULOS))
    concorrentes.update(_ler_blocos_ano_a_ano(_get_rows(wb_curitiba, "2025x2026"), CURITIBA_ROTULOS))
    concorrentes.update(_ler_blocos_foz(_get_rows(wb_foz, "2024x2025X2026"), FOZ_ROTULOS))

    reguas_out = []
    for regua in REGUAS:
        mensal = {}
        for m, mes in enumerate(MESES_PT):
            mes_dado = visitacao_por_mes.get(mes.upper())
            entry = {}
            for parque in regua["nossoParques"]:
                v = None
                if mes_dado:
                    v = (mes_dado.get("summary", {}).get(parque, {}) or {}).get("realizado")
                entry[parque] = v
            for concorrente in regua["concorrentes"]:
                valores = concorrentes.get(concorrente)
                entry[concorrente] = valores[m] if valores else None
            mensal[mes] = entry
        reguas_out.append({
            "id": regua["id"], "regiao": regua["regiao"], "nome": regua["nome"],
            "nossoParques": regua["nossoParques"], "concorrentes": regua["concorrentes"],
            "mensal": mensal,
        })

    return {"reguas": reguas_out}


# =========================================================================================
# COMO INTEGRAR no extract_data.py real
# =========================================================================================
# 1. Salvar este arquivo como scripts/intel_mercado_extension.py (mesma pasta do
#    extract_data.py) + import no topo do extract_data.py:
#      from intel_mercado_extension import build_intel_mercado
# 2. config.json (mesma pasta) precisa ter a chave nova:
#      "intel_mercado_ids": {
#        "rio": "1aaEQrLUwbRZakkKJSdjLc-S9Wpa3U--WJCDTb1B-k_w",
#        "aparecida": "13FG91m9kZiZTlYqWospg5HeSf2v8Mf3C9KBSwN9nMYI",
#        "curitiba": "1rbYIehrdmg4wJRmM9q_a8PEKM1fBLVMZ3JyRBeCzcFg",
#        "foz": "1dbfeeQV5bL9utmuJaXp3bQTbFcExjswxkcaw5eDOQHg"
#      }
# 3. Em main(), depois que output["VISITACAO"] já estiver montado, dentro de um try/except
#    (nunca deixar essa aba derrubar o resto do pipeline -- mesmo padrão das outras seções):
#      try:
#          output["INTEL_MERCADO"] = build_intel_mercado(
#              service, cfg["intel_mercado_ids"], output["VISITACAO"], _bioparque_ainda_conta
#          )
#      except Exception as e:
#          print(f"AVISO: falha ao montar Inteligência de Mercado: {e}", file=sys.stderr)
#          output["INTEL_MERCADO"] = {"reguas": []}
# 4. As 4 planilhas (Rio/Aparecida/Curitiba/Foz) precisam estar compartilhadas com a service
#    account do pipeline (senão a leitura cai no except e a aba fica vazia -- ver aviso no log
#    da Action).
# =========================================================================================
