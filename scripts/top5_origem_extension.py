# -*- coding: utf-8 -*-
"""
Extensão do extract_data.py: TOP 5 país (Grupo Cataratas) / estado (Soul) — Mix de Origem
============================================================================================

CONTEXTO
--------
Pedido do usuário (19/08/2026): mostrar, em cada card da aba "Mix de Origem" do painel, um
mini-ranking dos 5 países (parques Grupo Cataratas) ou estados (parques Soul: Três Pescadores
e Vila Velha) de onde vêm os visitantes, com medalha ouro/prata/bronze nos 3 primeiros. O
front-end (index.html) já está pronto e espera uma nova chave `MIX_ORIGEM_TOP5` no data.json,
no formato:

    {
      "<Parque>": {
        "tipo": "pais" | "estado",
        "daily": { "YYYY-MM-DD": { "<País ou UF>": <qtd int>, ... }, ... },
        "total": <int>   # soma de todos os dias/labels -- não usado pelo front-end hoje,
                          # mas útil pra debug/conferência
      },
      ...
    }

Esse arquivo (ou o conteúdo dele colado dentro de extract_data.py) monta essa estrutura a
partir dos EXPORTS BRUTOS de bilheteria por parque (um arquivo .xlsx por mês por parque, com
colunas PAIS/UF por venda) que vivem numa pasta do Google Drive -- NÃO da aba "ANÁLISE MIX DE
ORIGEM 2026" (que só tem Local/Brasileiro/Estrangeiro, sem quebra de país/estado).

Este módulo processou e validou manualmente Agosto/2026 (rodando fora do pipeline, num
ambiente sandbox sem acesso de push ao repositório) -- ver `topmix/aggregate_top5.py` e o
`top5_agosto.json` gerado por ele, já mesclados no data.json entregue junto com o index.html
novo. Para os próximos meses (e pra manter isso rodando sozinho todo dia, sem alguém ter que
rodar isso manualmente de novo), esse código precisa ser colado no extract_data.py real e
integrado ao pipeline -- ver "COMO INTEGRAR" no fim do arquivo.


SCHEMAS DOS ARQUIVOS BRUTOS (conferidos manualmente em Agosto/2026)
--------------------------------------------------------------------
Grupo Cataratas (país-ranked) -- AquaRio, BioParque, Paineiras, M3F, AquaFoz:
    Uma pasta por parque no Drive, um .xlsx por mês. Colunas relevantes:
      PARQUE_ORIGEM   -- o arquivo pode misturar linhas de mais de um parque; SEMPRE filtrar
                         por este campo (valor em maiúsculas: AQUARIO, BIOPARQUE, PAINEIRAS,
                         M3F, AQUAFOZ)
      DATA_VISITACAO  -- pode conter datas fora do mês (ex.: ingresso vendido antecipado pra
                         mês seguinte) -- SEMPRE filtrar por ano/mês também
      PAIS            -- pode vir em branco (ingressos sem nacionalidade registrada, ex.:
                         categoria "Guia de Turismo" -- ~1 a 22% dos tickets dependendo do
                         parque). Linhas sem PAIS são corretamente excluídas do ranking (não
                         tem como saber de onde vieram) -- NÃO tentar inferir a partir de UF.
      QTDE_TOTAL      -- quantidade de ingressos daquela linha

    PNI é um caso à parte dentro do grupo Cataratas: tem arquivo próprio ("Visitação UF ate
    DD.MM.2026.xlsx"), aba "Planilha3" (a "Planilha1" é só uma tabela dinâmica-resumo, não
    usar), colunas: DATA.1 (data completa), ANO, NMES (mês numérico), NOME_PAIS, ESTADO,
    QUANTIDADE. PNI é ranqueado por NOME_PAIS (mesmo critério dos demais Cataratas).

Soul (estado-ranked) -- Três Pescadores, Vila Velha:
    Mesma lógica de pasta+arquivo mensal, mas schema diferente: SEM coluna PARQUE_ORIGEM (cada
    arquivo já é de um único parque), coluna de quantidade chamada QUANTIDADE_REALIZADA (não
    QTDE_TOTAL), colunas PAIS/UF (UF é o que usamos pro ranking; "EX"/"EXTERIOR" em UF vira o
    bucket "Exterior"). Três Pescadores tem coluna extra MUNICIPIO, não usada aqui.


CORREÇÃO DE PAÍS CORROMPIDO/VERBOSO (PAIS_FIXUP)
--------------------------------------------------
Os exports do grupo Cataratas têm uma corrupção de encoding real: caracteres acentuados em
alguns nomes de país viraram o caractere de substituição "�" (ex.: "Fran�a" em vez de
"França"). Sem correção, o norm() (que só remove acento) descarta o "�" e produz "Frana" em
vez de "Franca" -- um país real com nome errado no ranking. O PNI não tem essa corrupção, mas
usa nomes burocráticos longos ("Armenia, República da") que não batem com a forma curta usada
nos demais parques ("Armênia"). PAIS_FIXUP mapeia toda variante bruta conhecida (corrompida,
burocrática ou grafia alternativa) pra uma forma canônica única, ANTES do norm() cuidar de
acento/caixa. A lista abaixo foi construída a partir da união de TODOS os valores únicos de
PAIS/NOME_PAIS realmente encontrados nos 6 arquivos país-ranked de Agosto/2026 -- ao rodar
outros meses, é bem possível aparecer alguma variante nova que não está aqui ainda; se algum
país estranho aparecer no Top 5 (nome truncado, cortado por um "�" isolado, etc.), é sinal de
que precisa entrar uma entrada nova neste dicionário.
"""

import unicodedata

PAIS_FIXUP = {
    # --- mojibake (caractere de substituição) -> forma correta acentuada ---
    'Afeganist�o': 'Afeganistão',
    'Alb�nia, Rep�blica da': 'Albânia',
    'Arg�lia': 'Argélia',
    'Armenia, Rep�blica da': 'Armênia',
    'Ar�bia Saudita': 'Arábia Saudita',
    'Canad�': 'Canadá',
    'Cazaquist�o, Rep�blica do': 'Cazaquistão',
    'Croacia, Rep�blica da': 'Croácia',
    'Eslovaca, Rep�blica': 'Eslováquia',
    'Eslov�nia, Rep�blica da': 'Eslovênia',
    'Est�nia, Rep�blica da': 'Estônia',
    'Fran�a': 'França',
    'Gales, Pa�s de': 'País de Gales',
    'Georgia, Rep�blica da': 'Geórgia',
    'Groenl�ndia': 'Groenlândia',
    'Hong Kong, Regi�o Adm. Especial': 'Hong Kong',
    'Holanda (Pa�ses Baixos)': 'Holanda',
    'Hungria, Rep�blica da': 'Hungria',
    'Indon�sia': 'Indonésia',
    'Letonia, Rep�blica da': 'Letônia',
    'Lituania, Rep�blica da': 'Lituânia',
    'Mold�via, Rep�blica da': 'Moldávia',
    'Mong�lia': 'Mongólia',
    'Mo�ambique': 'Moçambique',
    'Nicar�gua': 'Nicarágua',
    'Nig�ria': 'Nigéria',
    'Nova Caled�nia': 'Nova Caledônia',
    'Nova Zel�ndia': 'Nova Zelândia',
    'Om�': 'Omã',
    'Panam�': 'Panamá',
    'Papua Nova Guin�': 'Papua Nova Guiné',
    'Pol�nia, Rep�blica da': 'Polônia',
    'Rep�blica Dominicana': 'República Dominicana',
    'Rom�nia': 'Romênia',
    'Sao Tom� e Pr�ncipe, Ilhas': 'São Tomé e Príncipe',
    'Siria, Rep�blica �rabe da': 'Síria',
    'Su�cia': 'Suécia',
    'Su��a': 'Suíça',
    'Tail�ndia': 'Tailândia',
    'Vietn�': 'Vietnã',
    'Zimb�bue': 'Zimbábue',
    # --- formas burocráticas/verbosas (sem corrupção) -> forma curta ---
    'Albânia, República da': 'Albânia',
    'Armenia, República da': 'Armênia',
    'Azerbaijao, República do': 'Azerbaijão',
    'Bahamas, Ilhas': 'Bahamas',
    'Bahrein, Ilhas': 'Barém',
    'Cayman, Ilhas': 'Ilhas Cayman',
    'Cazaquistão, República do': 'Cazaquistão',
    'Comores, Ilhas': 'Comores',
    'Congo, República Democrática do': 'Congo',
    'Croacia, República da': 'Croácia',
    'Dominica, Ilha': 'Dominica',
    'Eslovaca, República': 'Eslováquia',
    'Eslovênia, República da': 'Eslovênia',
    'Estônia, República da': 'Estônia',
    'Gales, País de': 'País de Gales',
    'Georgia, República da': 'Geórgia',
    'Hong Kong, Região Adm. Especial': 'Hong Kong',
    'Holanda (Países Baixos)': 'Holanda',
    'Hungria, República da': 'Hungria',
    'Ira, República Islâmica do': 'Irã',
    'Letonia, República da': 'Letônia',
    'Lituania, República da': 'Lituânia',
    'Man, Ilhas': 'Ilha de Man',
    'Marshall, Ilhas': 'Ilhas Marshall',
    'Midway, Ilhas': 'Ilhas Midway',
    'Moldávia, República da': 'Moldávia',
    'Norfolk, Ilha': 'Ilha Norfolk',
    'Polônia, República da': 'Polônia',
    'Republica Dominicana': 'República Dominicana',
    'Reunião, Ilha': 'Ilha Reunião',
    'Siria, República Árabe da': 'Síria',
    'Tanzânia, República Unida da': 'Tanzânia',
    'Uzbequistão, República do': 'Uzbequistão',
    'Vaticano, Estado da Cidade do': 'Vaticano',
    'Virgens, Ilhas (E.U.A.)': 'Ilhas Virgens Americanas',
    'Wallis e Futuna, Ilhas': 'Wallis e Futuna',
    'Christmas, Ilha (Navidad)': 'Ilha Christmas',
    # --- grafias alternativas (não é só acento) ---
    'Antiga e Barbuda': 'Antígua e Barbuda',
    'Antigua e Barbuda': 'Antígua e Barbuda',
    'Bermuda': 'Bermudas',
    'Burquina Faso': 'Burkina Faso',
    'Curacao': 'Curaçao',
    'Federacao Russa': 'Rússia',
    'Quirquistao': 'Quirguistão',
    'Outros Paises': 'Outros',
    'Belarus': 'Bielorrússia',
}


def norm(s):
    """Normaliza um valor bruto de PAIS/NOME_PAIS pra uma forma canônica única: aplica
    PAIS_FIXUP (corrupção/verbosidade), depois remove acento e padroniza caixa (Title Case)
    pra não duplicar "Brasil" vs "BRASIL" vs "brasil" como países diferentes no ranking."""
    if s is None:
        return None
    s = str(s).strip()
    if not s or s.lower() in ('nan', 'none'):
        return None
    s = PAIS_FIXUP.get(s, s)
    s2 = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('ascii')
    return s2.strip().title()


def uf_norm(v):
    """Normaliza UF (parques Soul): mantém a sigla, mas unifica EX/EXTERIOR num único
    bucket "Exterior" (mais legível no ranking que a sigla crua)."""
    if v is None:
        return None
    s = str(v).strip().upper()
    if not s or s.lower() in ('nan', 'none'):
        return None
    if s in ('EX', 'EXTERIOR'):
        return 'Exterior'
    return s


# ---------------------------------------------------------------------------------------
# Config: schema de cada grupo de parque. `origem` é o valor de PARQUE_ORIGEM (Cataratas) que
# identifica as linhas daquele parque dentro do arquivo (que pode conter mais de um parque).
# ---------------------------------------------------------------------------------------
CATARATAS_PARKS = {
    'AquaRio':   dict(origem='AQUARIO'),
    'BioParque': dict(origem='BIOPARQUE'),
    'Paineiras': dict(origem='PAINEIRAS'),
    'M3F':       dict(origem='M3F'),
    'AquaFoz':   dict(origem='AQUAFOZ'),
}
SOUL_PARKS = ['Três Pescadores', 'Vila Velha']
# PNI processado à parte (process_pni_sheet) por ter schema próprio.


def _header_index(header_row):
    """Mapa {nome_da_coluna: índice} a partir da 1ª linha de uma planilha (aba Planilha3 do
    PNI ou qualquer export bruto de bilheteria) -- todos com cabeçalho na 1ª linha."""
    return {str(v).strip(): i for i, v in enumerate(header_row) if v is not None}


def _iter_rows(ws):
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = _header_index(header)
    for row in rows:
        yield row, idx


def process_cataratas_sheet(ws, origem, ano, mes_num):
    """Processa um export bruto de um parque do Grupo Cataratas (exceto PNI). Devolve
    (daily_dict, total_com_pais) -- daily_dict no formato {'YYYY-MM-DD': {pais: qtd}}."""
    daily = {}
    total = 0
    for row, idx in _iter_rows(ws):
        if str(row[idx['PARQUE_ORIGEM']] or '').strip().upper() != origem:
            continue
        data_visita = row[idx['DATA_VISITACAO']]
        if data_visita is None:
            continue
        # data_visita pode vir como datetime (openpyxl converte automaticamente células de
        # data) -- se vier como string, adaptar o parse aqui.
        if getattr(data_visita, 'year', None) != ano or getattr(data_visita, 'month', None) != mes_num:
            continue
        pais = norm(row[idx['PAIS']])
        if pais is None:
            continue  # sem PAIS registrado (ex.: Guia de Turismo) -- não entra no ranking
        qtd = row[idx['QTDE_TOTAL']] or 0
        dia = data_visita.strftime('%Y-%m-%d')
        daily.setdefault(dia, {})
        daily[dia][pais] = daily[dia].get(pais, 0) + int(qtd)
        total += int(qtd)
    return daily, total


def process_soul_sheet(ws, ano, mes_num):
    """Processa um export bruto de um parque Soul (Três Pescadores/Vila Velha). Devolve
    (daily_dict, total_com_uf) -- daily_dict no formato {'YYYY-MM-DD': {uf: qtd}}."""
    daily = {}
    total = 0
    for row, idx in _iter_rows(ws):
        data_visita = row[idx['DATA_VISITACAO']]
        if data_visita is None:
            continue
        if getattr(data_visita, 'year', None) != ano or getattr(data_visita, 'month', None) != mes_num:
            continue
        uf = uf_norm(row[idx['UF']])
        if uf is None:
            continue  # sem UF registrado (ex.: linhas de combo sem venda avulsa) -- fora do ranking
        qtd = row[idx['QUANTIDADE_REALIZADA']] or 0
        dia = data_visita.strftime('%Y-%m-%d')
        daily.setdefault(dia, {})
        daily[dia][uf] = daily[dia].get(uf, 0) + int(qtd)
        total += int(qtd)
    return daily, total


def process_pni_sheet(ws, ano, mes_num):
    """Processa a aba 'Planilha3' do arquivo de PNI (NÃO usar 'Planilha1', que é só um
    resumo/tabela dinâmica). Devolve (daily_dict, total_com_pais)."""
    daily = {}
    total = 0
    for row, idx in _iter_rows(ws):
        if row[idx['ANO']] != ano or row[idx['NMES']] != mes_num:
            continue
        pais = norm(row[idx['NOME_PAIS']])
        if pais is None:
            continue
        qtd = row[idx['QUANTIDADE']] or 0
        data_completa = row[idx['DATA.1']]
        dia = data_completa.strftime('%Y-%m-%d') if data_completa else f"{ano}-{mes_num:02d}-01"
        daily.setdefault(dia, {})
        daily[dia][pais] = daily[dia].get(pais, 0) + int(qtd)
        total += int(qtd)
    return daily, total


def _merge_daily(alvo, novo):
    """Mescla um novo daily_dict (de um mês) dentro do acumulado de todos os meses já
    processados pro mesmo parque, somando quando o mesmo dia aparece 2x (não deveria
    acontecer se cada mês só é processado 1x, mas é seguro contra reprocessamento)."""
    for dia, porLabel in novo.items():
        alvo.setdefault(dia, {})
        for label, qtd in porLabel.items():
            alvo[dia][label] = alvo[dia].get(label, 0) + qtd


def escolher_melhor_arquivo(drive_service, folder_id, ano, mes_num, ler_e_somar_fn, reference_total=None):
    """Quando uma pasta do Drive tem MAIS DE UM arquivo candidato pro mesmo mês (duplicidade
    -- situação real encontrada em Agosto/2026 pra Paineiras), baixa e processa CADA
    candidato e escolhe o melhor, nesta ordem de prioridade:

      1) Se `reference_total` foi passado (ex.: o "Total Mês"/"Total Período" já calculado
         pela leitura oficial de MIX_ORIGEM/MIX_ORIGEM_DIARIO pro mesmo parque/mês): o
         candidato cujo total somado fica mais PRÓXIMO do valor de referência vence. Método
         usado manualmente em Agosto/2026 pra decidir entre 2 arquivos de Paineiras (um
         somava 63.998, batendo com o Total Mês oficial de 64.637 -- ~1% de variação
         aceitável por causa do instante do snapshot -- o outro só ia até 04/08, claramente
         desatualizado).
      2) Sem referência (ou empate): o arquivo com `modifiedTime` mais recente vence.

    ler_e_somar_fn(wb) -> (daily_dict, total) é uma das process_*_sheet acima (já parcialmente
    aplicada com ano/mes_num/origem via functools.partial ou lambda, ver build_mix_origem_top5).
    Devolve (daily_dict, total) do arquivo escolhido, ou (None, None) se a pasta não tiver
    nenhum arquivo com dado pro mês pedido.
    """
    import functools

    resp = drive_service.files().list(
        q=f"'{folder_id}' in parents and trashed=false",
        fields="files(id,name,modifiedTime,mimeType)",
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    candidatos = resp.get('files', [])
    if not candidatos:
        return None, None

    resultados = []  # (daily, total, modifiedTime)
    for f in candidatos:
        wb = _download_workbook_generic(drive_service, f['id'], f['mimeType'])
        try:
            daily, total = ler_e_somar_fn(wb)
        except Exception:
            continue  # arquivo sem os dados esperados (aba errada, schema diferente) -- pula
        if daily:  # só considera candidatos que realmente têm linhas do mês pedido
            resultados.append((daily, total, f['modifiedTime']))

    if not resultados:
        return None, None
    if len(resultados) == 1:
        return resultados[0][0], resultados[0][1]

    if reference_total:
        resultados.sort(key=lambda r: abs(r[1] - reference_total))
        return resultados[0][0], resultados[0][1]

    resultados.sort(key=lambda r: r[2], reverse=True)  # modifiedTime mais recente primeiro
    return resultados[0][0], resultados[0][1]


def _download_workbook_generic(drive_service, file_id, mime_type):
    """Igual a _download_workbook() do extract_data.py, mas recebe o mimeType já conhecido
    (evita 1 chamada extra files().get()) -- reaproveitar a função original do
    extract_data.py se for colar este código dentro dele (ver COMO INTEGRAR)."""
    import io
    import openpyxl
    from googleapiclient.http import MediaIoBaseDownload

    buf = io.BytesIO()
    if mime_type == "application/vnd.google-apps.spreadsheet":
        request = drive_service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(buf, request, chunksize=10 * 1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=5)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True, read_only=True)


def build_mix_origem_top5(drive_service, top5_folders, meses_com_dados, ano=2026, reference_totals=None):
    """Função principal -- chamar de dentro de main() do extract_data.py e salvar o
    resultado em data['MIX_ORIGEM_TOP5'].

    top5_folders: dict {parque: drive_folder_id} -- UMA pasta do Drive por parque (a mesma
        pasta que o usuário compartilhou manualmente, com um .xlsx por mês dentro). Precisa
        ser adicionado ao config.json (ex.: chave "top5_origem_folders") -- os IDs de pasta
        (não de arquivo!) usados nesta sessão sandbox eram de uma pasta por parque dentro de
        https://drive.google.com/drive/folders/168VZDzJtidrcN2z54LCuJTgHjjTRPqZB -- usar
        drive_service.files().list() com esse ID como parents pra descobrir as subpastas de
        cada parque, se ainda não tiverem sido copiadas pro config.json.
    meses_com_dados: lista de (mes_num, mes_nome) a processar -- tipicamente
        range(1, mes_atual+1), ou reaproveitar config["meses_com_dados"] convertido pra número.
    reference_totals: dict opcional {(parque, mes_num): total_oficial_do_mes} -- usado só
        pra resolver duplicidade de arquivo (ver escolher_melhor_arquivo). Pode vir do
        MIX_ORIGEM_DIARIO já calculado neste mesmo run do pipeline (soma de v26 do mês).
        Sem essa referência, duplicidade é resolvida por modifiedTime mais recente.

    Devolve o dict pronto pra virar data['MIX_ORIGEM_TOP5'] (ver formato no topo do arquivo).
    """
    import functools

    result = {}

    for parque, cfg in CATARATAS_PARKS.items():
        folder_id = top5_folders.get(parque)
        if not folder_id:
            continue
        daily_total = {}
        for mes_num, _ in meses_com_dados:
            ref = (reference_totals or {}).get((parque, mes_num))
            fn = functools.partial(lambda wb, o=cfg['origem'], m=mes_num: process_cataratas_sheet(wb.active, o, ano, m))
            daily, _total = escolher_melhor_arquivo(drive_service, folder_id, ano, mes_num, fn, ref)
            if daily:
                _merge_daily(daily_total, daily)
        if daily_total:
            total = sum(sum(d.values()) for d in daily_total.values())
            result[parque] = dict(tipo='pais', daily=daily_total, total=total)

    # PNI: schema próprio, aba Planilha3 -- mesma pasta/arquivo-por-mês, mas quem processa é
    # process_pni_sheet, e o total de referência (se houver) sai da mesma fonte que os demais.
    pni_folder = top5_folders.get('PNI')
    if pni_folder:
        daily_total = {}
        for mes_num, _ in meses_com_dados:
            ref = (reference_totals or {}).get(('PNI', mes_num))
            fn = functools.partial(lambda wb, m=mes_num: process_pni_sheet(wb['Planilha3'], ano, m))
            daily, _total = escolher_melhor_arquivo(drive_service, pni_folder, ano, mes_num, fn, ref)
            if daily:
                _merge_daily(daily_total, daily)
        if daily_total:
            total = sum(sum(d.values()) for d in daily_total.values())
            result['PNI'] = dict(tipo='pais', daily=daily_total, total=total)

    for parque in SOUL_PARKS:
        folder_id = top5_folders.get(parque)
        if not folder_id:
            continue
        daily_total = {}
        for mes_num, _ in meses_com_dados:
            ref = (reference_totals or {}).get((parque, mes_num))
            fn = functools.partial(lambda wb, m=mes_num: process_soul_sheet(wb.active, ano, m))
            daily, _total = escolher_melhor_arquivo(drive_service, folder_id, ano, mes_num, fn, ref)
            if daily:
                _merge_daily(daily_total, daily)
        if daily_total:
            total = sum(sum(d.values()) for d in daily_total.values())
            result[parque] = dict(tipo='estado', daily=daily_total, total=total)

    return result


# =========================================================================================
# COMO INTEGRAR no extract_data.py real
# =========================================================================================
# 1. Colar as constantes/funções deste arquivo dentro de extract_data.py (ou importar este
#    módulo, se preferir manter separado -- nesse caso adicionar "top5_origem_extension.py"
#    ao mesmo diretório scripts/ e fazer `from top5_origem_extension import build_mix_origem_top5`).
#
# 2. Adicionar ao config.json uma chave nova com o ID de pasta (não de arquivo!) do Drive de
#    cada parque, ex.:
#      "top5_origem_folders": {
#        "AquaRio": "<folder_id>", "BioParque": "<folder_id>", "Paineiras": "<folder_id>",
#        "M3F": "<folder_id>", "AquaFoz": "<folder_id>", "PNI": "<folder_id>",
#        "Três Pescadores": "<folder_id>", "Vila Velha": "<folder_id>"
#      }
#    (a pasta raiz compartilhada pelo usuário nesta sessão tem uma subpasta por parque --
#    usar drive_service.files().list(q="'<ID_DA_PASTA_RAIZ>' in parents") pra listar e pegar
#    o ID de cada uma.)
#
# 3. Em main(), depois que MIX_ORIGEM_DIARIO já estiver calculado (pra poder usar como
#    reference_totals opcional -- ver função build_mix_origem_diario existente), chamar:
#
#      top5_folders = config.get("top5_origem_folders", {})
#      if top5_folders:
#          meses_num = [(i+1, m) for i, m in enumerate(config["meses"]) if m in config["meses_com_dados"]]
#          data["MIX_ORIGEM_TOP5"] = build_mix_origem_top5(service, top5_folders, meses_num)
#      else:
#          data["MIX_ORIGEM_TOP5"] = {}  # pasta ainda não configurada -- front-end trata vazio normalmente
#
# 4. Testar localmente (fora do GitHub Action) antes de confiar no cron: rodar
#    `python scripts/extract_data.py --config scripts/config.json --out /tmp/data_teste.json`
#    e conferir se `MIX_ORIGEM_TOP5` saiu com os 8 parques e números plausíveis (comparar o
#    total de um parque/mês com o "Total Mês" que já aparece na tabela normal do Mix de
#    Origem daquele parque -- uma diferença de até ~5-20% é esperada e normal, por causa de
#    ingressos sem PAIS/UF registrado -- ex.: categoria "Guia de Turismo" -- que ficam de
#    fora do ranking por país/estado mas entram no total geral do parque).
#
# 5. ATENÇÃO -- arquivos grandes: alguns exports mensais têm 12-27MB e algumas pastas têm até
#    5 candidatos duplicados pro mesmo mês (escolher_melhor_arquivo baixa e processa TODOS os
#    candidatos pra decidir) -- rodar o histórico completo (Jan-Ago) de uma vez pode ser lento
#    e pesado a primeira vez que isso rodar no GitHub Action. Considere: (a) rodar o backfill
#    completo uma única vez manualmente/localmente e colar o resultado direto num
#    MIX_ORIGEM_TOP5_HISTORICO fixo, deixando o cron diário só processar o MÊS ABERTO daí em
#    diante: ou (b) aumentar o timeout do job do GitHub Action na primeira execução.
# =========================================================================================
