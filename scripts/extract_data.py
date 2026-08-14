#!/usr/bin/env python3
"""
Extrai os dados das 4 planilhas do Grupo Cataratas e gera um data.json consolidado
para o painel_parques_rio.html.

A logica de leitura de cada aba foi validada manualmente (celula a celula) durante a
auditoria de dados desta sessao -- inclusive os dois bugs encontrados (Museu de Cera
zerado em Marco/Maio por causa de uma linha duplicada vazia, e Captacao CV em dobro por
somar a linha de total junto com os dias). O codigo abaixo evita os dois de proposito;
os comentarios marcados com "# BUG EVITADO:" explicam onde.

Uso:
    export GOOGLE_APPLICATION_CREDENTIALS=/caminho/para/service-account.json
    python extract_data.py --config config.json --out ../data.json

Requer: google-api-python-client, google-auth, openpyxl (ver requirements.txt)

NOTA IMPORTANTE: as planilhas do Grupo Cataratas estao guardadas no Drive como arquivos
.xlsx (Excel de verdade), nao como Google Sheets nativos. A API do Google Sheets nao le
arquivos .xlsx diretamente ("This operation is not supported for this document. The
document must not be an Office file."). Por isso este script usa a API do **Google Drive**
para baixar o arquivo (funciona tanto pra .xlsx quanto pra Google Sheets nativo) e le o
conteudo com openpyxl -- a mesma biblioteca usada pra validar os dados nesta conversa.
"""
import argparse
import calendar
import datetime
import io
import json
import re
import socket
import sys
import unicodedata

import openpyxl
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# O runner do GitHub Actions as vezes tem uma conexao mais lenta/instavel com a API do
# Google, e o timeout padrao do socket (ilimitado, mas o cliente HTTP interno usa um valor
# curto) estourava no meio do download de planilhas maiores. Aumenta a margem e deixa o
# retry automático (num_retries em next_chunk, abaixo) lidar com falhas passageiras.
socket.setdefaulttimeout(300)

SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]

MONTH_NUMBER = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARÇO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
    "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

PARKS = ["AquaRio", "BioParque", "Paineiras", "PNI", "M3F", "AquaFoz", "Três Pescadores", "Vila Velha"]
ATRATIVOS = ["GEX", "MDE", "MDC"]

# BioParque saiu do Grupo Cataratas a partir de Agosto/2026 (pedido do usuário, 06/08/2026).
# Meses fechados (Jan-Jul/2026) e todo o histórico anterior continuam contando pro
# acumulado/histórico normalmente -- só a partir de Agosto/2026 o parque precisa
# desaparecer de toda visão "atual" do painel (linha da tabela, réguas, dropdowns,
# agregados do Grupo Cataratas, Report, Mix de Origem, Share, Investimento, Plano de
# Mídia). Um único ponto de verdade aqui evita ter que caçar essa regra em 10 funções
# diferentes -- cada builder chama uma das duas funções abaixo, do jeito que combinar
# com o formato de "mês" que ele já usa (nome PT completo, "AGOSTO", ou (ano, mes_num)).
BIOPARQUE_SAIDA_ANO_MES = (2026, 8)  # (ano, mes numérico) do 1º período que NÃO conta mais
BIOPARQUE_MESES_PT_POS_SAIDA = {
    "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
}
BIOPARQUE_MESES_PT_MAIUSCULO_POS_SAIDA = {m.upper() for m in BIOPARQUE_MESES_PT_POS_SAIDA}


def _bioparque_ainda_conta(ano, mes_numero):
    """True se BioParque ainda contava como Grupo Cataratas nesse ano/mês (todo o
    histórico até Julho/2026); False a partir de Agosto/2026."""
    return (ano, mes_numero) < BIOPARQUE_SAIDA_ANO_MES

SHEETS_EPOCH = datetime.date(1899, 12, 30)


def serial_to_date(serial):
    """Converte o numero serial de data (mesmo formato usado pela API do Sheets) para
    datetime.date. Mantido por compatibilidade -- get_values() ja devolve as datas nesse
    formato numerico, igual antes."""
    if serial is None:
        return None
    try:
        return SHEETS_EPOCH + datetime.timedelta(days=float(serial))
    except (TypeError, ValueError):
        return None


def get_drive_service():
    creds = service_account.Credentials.from_service_account_file(
        _sa_key_path(), scopes=SCOPES
    ) if _sa_key_path() else service_account.Credentials.from_service_account_info(
        json.loads(_sa_key_env()), scopes=SCOPES
    )
    return build("drive", "v3", credentials=creds)


def _sa_key_path():
    import os
    return os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")


def _sa_key_env():
    import os
    # alternativa: colar o JSON inteiro direto numa env var (usado no workflow do GitHub)
    return os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "{}")


_WORKBOOK_CACHE = {}


def _download_workbook(drive_service, spreadsheet_id):
    """Baixa o arquivo (Drive API) e abre com openpyxl. Cacheia por spreadsheet_id pra nao
    baixar o mesmo arquivo de novo a cada aba lida (ex.: 7 abas de mes na mesma planilha)."""
    if spreadsheet_id in _WORKBOOK_CACHE:
        return _WORKBOOK_CACHE[spreadsheet_id]

    # supportsAllDrives=True e' necessario quando o arquivo vive dentro de um Drive
    # Compartilhado (Shared Drive) da empresa -- sem isso a API responde "File not found"
    # mesmo quando o arquivo esta compartilhado corretamente com a service account.
    meta = drive_service.files().get(
        fileId=spreadsheet_id, fields="mimeType, name", supportsAllDrives=True
    ).execute()
    is_native_sheet = meta["mimeType"] == "application/vnd.google-apps.spreadsheet"

    buf = io.BytesIO()
    if is_native_sheet:
        # Google Sheets nativo: exporta como .xlsx antes de baixar
        request = drive_service.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        # ja e' um arquivo .xlsx (ou similar): baixa direto
        request = drive_service.files().get_media(fileId=spreadsheet_id, supportsAllDrives=True)

    # num_retries>0 faz o proprio googleapiclient re-tentar automaticamente (com backoff)
    # em caso de erro de rede passageiro (timeout, conexao resetada etc.) -- foi isso que
    # derrubou uma execucao no GitHub Actions ("TimeoutError: The read operation timed out").
    downloader = MediaIoBaseDownload(buf, request, chunksize=10 * 1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=5)
    buf.seek(0)

    wb = openpyxl.load_workbook(buf, data_only=True, read_only=True)
    _WORKBOOK_CACHE[spreadsheet_id] = wb
    return wb


def get_values(service, spreadsheet_id, sheet_name, a1_range=None):
    """Busca uma aba inteira como lista de listas. Datas viram numero serial (mesmo
    formato que a API do Sheets usava), pra nao precisar mudar o resto do codigo que
    espera esse formato."""
    wb = _download_workbook(service, spreadsheet_id)
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    width = max((len(r) for r in rows), default=0)
    out = []
    for r in rows:
        r = r + [None] * (width - len(r))
        r2 = []
        for v in r:
            if isinstance(v, datetime.datetime):
                r2.append(float((v.date() - SHEETS_EPOCH).days))
            else:
                r2.append(v)
        out.append(r2)
    return out


def cell(row, idx):
    return row[idx] if idx < len(row) else None


# ---------------------------------------------------------------------------
# Visitação Parques 2026.xlsx (uma aba por mês: JANEIRO..JULHO)
# ---------------------------------------------------------------------------

RESUMO_MENSAL_AGREGADOS = [
    "Grupo Cataratas (SSS)", "Grupo Cataratas SSS", "Total Grupo Cataratas",
    "Total Soul Parques", "Total Geral",
]


def _norm_label(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).strip().lower())


def _is_year(v, ano):
    try:
        return int(v) == ano
    except (TypeError, ValueError):
        return False


def _find_summary_columns(rows):
    """Acha as colunas de cada campo do resumo (mes E acumulado) pelo PROPRIO ROTULO na
    linha de cabecalho, em vez de confiar em offset fixo (1,2,3,4,5,7,8,9,10,11).

    BUG EVITADO: em Agosto/2026 o time inseriu uma coluna nova na aba (uma nota manual
    "Bio Até Julho" sobre a saida do BioParque), o que empurrou o bloco ACUMULADO inteiro
    uma coluna pra direita -- qualquer leitura por posição fixa desalinha TODOS os campos
    do acumulado (o que seria uma contagem grande passa a ser lido como se fosse um "%",
    virando números como 66916900,0%). Agora cada coluna e' achada pelo seu proprio rotulo
    na linha de cabecalho ("Realizado"/"OBZ"/"% OBZ"/2025/"% 2025" pro bloco do mes; 2026/
    "OBZ (Parcial)"/"%OBZ"/2025/"% 2025" pro bloco acumulado) -- funciona com qualquer
    coluna inserida/removida no meio, nos dois blocos.
    """
    header_idx = None
    for i, row in enumerate(rows[:6]):
        if any(_norm_label(c) == "realizado" for c in row):
            header_idx = i
            break
    if header_idx is None:
        return None
    header = rows[header_idx]
    acum_start = None
    for i, v in enumerate(header):
        if _is_year(v, 2026):
            acum_start = i
            break
    if acum_start is None:
        acum_start = len(header)

    cols = {}
    for i, v in enumerate(header):
        label = _norm_label(v)
        if i < acum_start:
            if label == "realizado" and "realizado" not in cols:
                cols["realizado"] = i
            elif label == "obz" and "obz" not in cols:
                cols["obz"] = i
            elif label == "% obz" and "pctObz" not in cols:
                cols["pctObz"] = i
            elif _is_year(v, 2025) and "y2025" not in cols:
                cols["y2025"] = i
            elif label == "% 2025" and "pct2025" not in cols:
                cols["pct2025"] = i
        else:
            if _is_year(v, 2026) and "acumRealizado" not in cols:
                cols["acumRealizado"] = i
            elif "%" in label and "obz" in label and "acumPctObz" not in cols:
                cols["acumPctObz"] = i
            elif "obz" in label and "acumObzParcial" not in cols:
                cols["acumObzParcial"] = i
            elif _is_year(v, 2025) and "acum2025" not in cols:
                cols["acum2025"] = i
            elif label == "% 2025" and "acumPct2025" not in cols:
                cols["acumPct2025"] = i
    return cols


def _read_label_rows(rows):
    """Acha a linha de cada parque/agregado do resumo mensal, na ORDEM em que aparecem na
    planilha -- usado por _detect_acum_row_shift pra saber quais linhas comparar."""
    found = []
    seen = set()
    for i, row in enumerate(rows[:20]):
        label = str(cell(row, 0) or "").strip()
        if not label:
            continue
        chave = None
        if label in RESUMO_MENSAL_AGREGADOS:
            chave = label
        else:
            for p in PARKS:
                if p.lower().replace("ê", "e") in label.lower().replace("ê", "e"):
                    chave = p
                    break
        if chave is None or chave in seen:
            continue
        seen.add(chave)
        found.append((i, chave))
    return found


def _detect_acum_row_shift(rows, label_rows, cols, bioparque_frozen=None):
    """Acha, com precisão, se o bloco ACUMULADO ficou desalinhado por causa de uma linha
    extra inserida (ex.: nota manual "Bio Até Julho" com o valor congelado do BioParque
    colado na linha de outro parque, empurrando o acumulado de todo mundo dali pra baixo
    uma posição, sem alterar o bloco do MES).

    BUG EVITADO: uma primeira versão desse detector tentava adivinhar o ponto de virada só
    com a matemática "Total Soul = 3P+VilaVelha / Total Geral = TotalGC+TotalSoul" -- mas
    essa conta fecha para VÁRIOS pontos de virada diferentes ao mesmo tempo (deslocar a
    partir do AquaRio, do Paineiras, da PNI... todos "fecham" a soma igualmente, porque
    deslocar um trecho contíguo em bloco preserva a soma interna dele mesmo estando errado
    linha a linha) -- então a conta sozinha NÃO identifica o ponto certo, só confirma que
    ALGO está deslocado. Em vez disso, usamos uma "impressão digital" única: já sabemos o
    valor exato do acumulado congelado do BioParque (calculado a partir da própria aba de
    Julho, antes dele saltar fora do Grupo Cataratas). Procuramos esse número EXATO nas
    linhas do bloco acumulado -- a linha onde ele aparece (colado por engano em cima do
    parque errado) é o ponto exato onde o deslocamento comeca, sem ambiguidade.
    """
    if not isinstance(bioparque_frozen, dict):
        return None
    alvo = bioparque_frozen.get("acumRealizado")
    if not isinstance(alvo, (int, float)):
        return None
    c_acum = cols.get("acumRealizado", 7)
    tol = max(1.0, abs(alvo) * 0.001)
    for i, _chave in label_rows:
        v = cell(rows[i], c_acum)
        if isinstance(v, (int, float)) and abs(v - alvo) <= tol:
            return i
    return None


def parse_month_summary(rows, bioparque_frozen=None):
    """Resumo por parque (linhas ~3-10, colunas do mes e do acumulado achadas por rotulo).

    BUG EVITADO (linhas): a versao anterior lia por POSICAO fixa (rows[2+i], uma linha por
    parque na ordem de PARKS) -- funcionava enquanto a aba tivesse exatamente as mesmas 8
    linhas de sempre, mas a partir de Agosto/2026 o time apagou a linha do BioParque na
    planilha ao vivo (ele saiu do Grupo Cataratas), o que empurrou Paineiras/PNI/M3F/
    AquaFoz/3P/Vila Velha uma linha pra cima -- e a leitura por posição passou a atribuir a
    cada parque o valor que era do PROXIMO. Agora cada linha e' identificada pelo seu
    proprio rotulo (coluna A), igual ja se faz em find_daily_blocks().

    BUG EVITADO (colunas): as colunas tambem sao achadas por rotulo (ver
    _find_summary_columns), porque uma coluna inserida no meio da aba quebraria a leitura
    por posição fixa igual as linhas quebravam.

    BUG EVITADO (linhas do ACUMULADO especificamente): mesmo com rotulo certo na linha do
    MES, o bloco ACUMULADO pode estar deslocado por causa de uma linha extra inserida SÓ
    naquele bloco (ver _detect_acum_row_shift) -- corrigido automaticamente procurando o
    valor exato do acumulado congelado do BioParque (parâmetro bioparque_frozen, vindo de
    Julho) dentro das linhas do acumulado; a linha onde ele aparece é o ponto exato do
    deslocamento.
    """
    cols = _find_summary_columns(rows) or {}
    c_realizado = cols.get("realizado", 1)
    c_obz = cols.get("obz", 2)
    c_pctObz = cols.get("pctObz", 3)
    c_y2025 = cols.get("y2025", 4)
    c_pct2025 = cols.get("pct2025", 5)
    c_acumRealizado = cols.get("acumRealizado", 7)
    c_acumObzParcial = cols.get("acumObzParcial", 8)
    c_acumPctObz = cols.get("acumPctObz", 9)
    c_acum2025 = cols.get("acum2025", 10)
    c_acumPct2025 = cols.get("acumPct2025", 11)

    label_rows = _read_label_rows(rows)
    shift_from = _detect_acum_row_shift(rows, label_rows, cols, bioparque_frozen)

    summary = {}
    for i, chave in label_rows:
        row = rows[i]
        acum_idx = i + 1 if (shift_from is not None and i >= shift_from) else i
        acum_row = rows[acum_idx] if acum_idx < len(rows) else row
        summary[chave] = {
            "realizado": cell(row, c_realizado),
            "obz": cell(row, c_obz),
            "pctObz": cell(row, c_pctObz),
            "y2025": cell(row, c_y2025),
            "pct2025": cell(row, c_pct2025),
            "acumRealizado": cell(acum_row, c_acumRealizado),
            "acumObzParcial": cell(acum_row, c_acumObzParcial),
            "acumPctObz": cell(acum_row, c_acumPctObz),
            "acum2025": cell(acum_row, c_acum2025),
            "acumPct2025": cell(acum_row, c_acumPct2025),
        }
    return summary


def find_daily_blocks(rows):
    """Acha, para cada parque, a linha 'Realizado 2026' e a linha 'OBZ 2026' logo abaixo.

    O titulo do bloco (nome do parque) fica 1 ou 2 linhas acima -- às vezes tem uma
    linha 'Ações' no meio, e o PNI aparece como 'URBIA + CATARATAS (PNI)' em vez de 'PNI'.
    """
    blocks = {}
    for i, row in enumerate(rows):
        if cell(row, 0) == "Realizado 2026":
            j = i - 1
            while j >= 0 and (not cell(rows[j], 0) or cell(rows[j], 0) == "Ações"):
                j -= 1
            title = str(cell(rows[j], 0) or "")
            norm = None
            for p in PARKS:
                if p.lower().replace("ê", "e") in title.lower().replace("ê", "e"):
                    norm = p
                    break
            if norm and norm not in blocks:
                blocks[norm] = i
    return blocks


def parse_month_daily(rows, n_days):
    daily = {}
    blocks = find_daily_blocks(rows)
    for park, ridx in blocks.items():
        realizado_row = rows[ridx][1:1 + n_days]
        obz_row = rows[ridx + 1][1:1 + n_days] if cell(rows[ridx + 1], 0) == "OBZ 2026" else [None] * n_days
        real2025_row = rows[ridx + 3][1:1 + n_days] if cell(rows[ridx + 3], 0) == "Realizado 2025" else [None] * n_days
        daily[park] = {
            "Realizado 2026": realizado_row,
            "OBZ 2026": obz_row,
            "Realizado 2025": real2025_row,
        }
    return daily


def parse_atrativos_daily(rows, n_days):
    """GEX/MDE/MDC diario.

    # BUG EVITADO: em Marco/Maio existem DUAS linhas 'MDC' na planilha (uma vazia, uma com
    # os dados de verdade, mais abaixo). Por isso sempre ficamos com a ULTIMA linha que
    # tiver algum valor não nulo -- nunca a primeira ocorrência do rótulo.
    """
    daily = {}
    for row in rows:
        label = cell(row, 0)
        if label in ATRATIVOS:
            vals = row[1:1 + n_days]
            if any(v is not None for v in vals):
                daily[label] = vals  # sobrescreve a anterior -> fica a ultima com dado
    return daily


def parse_atrativos_accum(rows):
    """Bloco lateral "ACUMULADO GEX | MDE | MDC" (por volta das linhas 3-5, colunas N em
    diante): nome, realizado2026, pctAq2026, realizado2025, pctAq2025.

    BUG EVITADO: a versao anterior lia por POSICAO fixa (coluna N = index 13 pro rotulo,
    colunas seguintes pros valores) -- quebrou quando a MESMA coluna extra que desalinhou o
    bloco ACUMULADO principal (ver _find_summary_columns) empurrou esse bloco lateral
    tambem uma coluna pra direita (rotulo "GEX"/"MDE"/"MDC" saiu da coluna N pra O). Agora
    o proprio rotulo e' procurado em cada linha, e os 4 valores são lidos relativos a onde
    ele foi encontrado -- funciona com qualquer deslocamento de coluna.
    """
    accum = {}
    for row in rows[:10]:
        for j, v in enumerate(row):
            if v in ATRATIVOS and v not in accum:
                accum[v] = {
                    "realizado2026": cell(row, j + 1),
                    "pctAq2026": cell(row, j + 2),
                    "realizado2025": cell(row, j + 3) if cell(row, j + 3) != "-" else None,
                    "pctAq2025": cell(row, j + 4) if cell(row, j + 4) != "-" else None,
                }
                break
    return accum


def detect_meses_com_dados(service, spreadsheet_id, meses_todos):
    """Descobre sozinho até qual mês a planilha 'Visitação Parques 2026.xlsx' já tem aba
    criada E com pelo menos um dia de "Realizado 2026" preenchido -- substitui a lista
    manual "meses_com_dados" do config.json, que antes precisava ser atualizada à mão
    todo mês assim que o resultado real do mês começava a entrar (o motivo de Agosto ter
    ficado só com Meta: a lista nunca foi atualizada).
    Para no primeiro mês em que: (a) a aba ainda não existe na planilha, ou (b) a aba já
    existe mas ainda está vazia (só o molde, sem nenhum "Realizado 2026" preenchido) --
    os meses seguintes na sequência também são tratados como sem dado, mesmo que por
    algum motivo tenham aba criada fora de ordem.
    BUG EVITADO: os meses futuros (ainda não iniciados) não ficam com a célula vazia --
    a planilha preenche "Realizado 2026" com 0 (não com célula em branco) até o mês
    realmente começar, igual já vimos nas abas DIÁRIO. Um teste de "existe valor não-nulo"
    conta esse 0 como "tem dado" e deixava passar o ano inteiro (Setembro a Dezembro
    incluídos, todos com 0). Por isso o teste exige um valor REALMENTE positivo (>0) em
    algum dia/parque, não só "diferente de None".
    """
    resultado = []
    for mes in meses_todos:
        try:
            rows = get_values(service, spreadsheet_id, mes)
        except Exception:
            break  # aba desse mês ainda não existe -- para aqui
        if not rows:
            break
        month_number = MONTH_NUMBER[mes]
        n_days = calendar.monthrange(2026, month_number)[1]
        daily = parse_month_daily(rows, n_days)
        tem_dado = any(
            any(isinstance(v, (int, float)) and v > 0 for v in (park_data.get("Realizado 2026") or []))
            for park_data in daily.values()
        )
        if not tem_dado:
            break
        resultado.append(mes)
    return resultado


def build_visitacao(service, spreadsheet_id, meses_com_dados):
    visitacao = {}
    # Snapshot do acumulado de Julho do BioParque -- usado para "congelar" o card dele na
    # visão ACUMULADO a partir de Agosto (pedido do usuário, 07/08/2026: "BioParque no
    # Acumulado pode constar, porém com valores até Julho com uma sinalização"). Guardamos
    # aqui porque a partir de Agosto a linha dele nem existe mais na planilha ao vivo.
    bioparque_julho_frozen = None
    for mes in meses_com_dados:
        rows = get_values(service, spreadsheet_id, mes)
        month_number = MONTH_NUMBER[mes]
        n_days = calendar.monthrange(2026, month_number)[1]
        summary = parse_month_summary(rows, bioparque_frozen=bioparque_julho_frozen)
        daily = parse_month_daily(rows, n_days)
        if _bioparque_ainda_conta(2026, month_number):
            if mes == "JULHO" and "BioParque" in summary:
                bioparque_julho_frozen = dict(summary["BioParque"])
        else:
            # BioParque saiu do Grupo Cataratas a partir de Agosto/2026 -- some da tabela
            # MENSAL/régua a partir daqui (Jan-Jul continuam intactos, lidos normalmente
            # nos ciclos anteriores deste mesmo loop). some de "daily" tambem (sem isso o
            # filtro do front `data.daily[park]` continuaria mostrando ele na tabela mensal).
            summary.pop("BioParque", None)
            daily.pop("BioParque", None)
            # No ACUMULADO, por pedido explícito do usuário, ele continua aparecendo -- mas
            # SÓ com o valor acumulado que já tinha fechado até Julho (congelado, nunca soma
            # nada de Agosto em diante), sinalizado no front via "historicoAteJulho".
            if bioparque_julho_frozen:
                summary["BioParque"] = {
                    "realizado": None, "obz": None, "pctObz": None,
                    "y2025": None, "pct2025": None,
                    "acumRealizado": bioparque_julho_frozen.get("acumRealizado"),
                    "acumObzParcial": bioparque_julho_frozen.get("acumObzParcial"),
                    "acumPctObz": bioparque_julho_frozen.get("acumPctObz"),
                    "acum2025": bioparque_julho_frozen.get("acum2025"),
                    "acumPct2025": bioparque_julho_frozen.get("acumPct2025"),
                    "historicoAteJulho": True,
                }
        visitacao[mes] = {
            "monthNumber": month_number,
            "nDays": n_days,
            "summary": summary,
            "daily": daily,
            "atrativos": {
                "daily": parse_atrativos_daily(rows, n_days),
                "accum": parse_atrativos_accum(rows),
            },
        }
    return visitacao


# ---------------------------------------------------------------------------
# Visitação Diária Ponderada - 2026 - Final.xlsx -> aba "Visitação diária GC V2"
# Tabela dinamica "Soma de Nova2026" (projecao diaria ponderada por parque, baseada no
# padrao de dias da semana de anos anteriores) -- usada como "Meta" nos meses que ainda
# nao tem resultado real (Agosto em diante). Por pedido do usuario, a Meta e' colocada no
# MESMO campo "OBZ 2026" que a regua diaria ja mostra (nao criamos um campo novo) -- assim
# a celula do dia mantem exatamente o layout de sempre (numero grande + "vs 25" + "OBZ"),
# só que com o numero grande e "vs 25" vazios ("-") e o "OBZ" preenchido com a meta.
# ---------------------------------------------------------------------------

VISITACAO_META_PARK_NAMES = {
    "Marco": "M3F", "3P": "Três Pescadores", "VV": "Vila Velha",
    "AquaFoz": "AquaFoz", "AquaRio": "AquaRio", "BioParque": "BioParque",
    "Paineiras": "Paineiras", "PNI": "PNI",
}

# BUG EVITADO: os "Rótulos de Coluna" dessa tabela dinamica NAO sao celulas de data de
# verdade (isinstance datetime) -- sao TEXTO no formato "01/mar", "02/mar" etc. (o Excel
# converte assim quando agrupa datas numa tabela dinamica). Por isso o parser abaixo le
# esse texto em vez de checar isinstance(..., datetime.datetime).
MES_ABREV_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def _parse_dd_mon(label, ano=2026):
    if not isinstance(label, str) or "/" not in label:
        return None
    dia_str, _, mes_str = label.strip().partition("/")
    mes_num = MES_ABREV_PT.get(mes_str.strip().lower())
    if mes_num is None or not dia_str.strip().isdigit():
        return None
    try:
        return datetime.date(ano, mes_num, int(dia_str))
    except ValueError:
        return None


# BUG EVITADO / TEMPORARIO: por pedido do usuario, os dados da planilha "Visitação
# Diária Ponderada - 2026" ainda NAO sao lidos ao vivo do Drive (a planilha ainda nao foi
# compartilhada com a service account) -- por enquanto, ficam "amarrados" aqui como dado
# ESTATICO, extraido uma unica vez da copia que o usuario enviou (Agosto a Dezembro/2026,
# projecao "Nova2026" por parque/dia). Quando a planilha for conectada ao Drive, trocar o
# uso de VISITACAO_META_ESTATICO em main() pela chamada a build_visitacao_meta() (a funcao
# ja esta pronta logo abaixo, so' nao esta sendo chamada por enquanto).
VISITACAO_META_ESTATICO = json.loads('''{"AGOSTO": {"Três Pescadores": [1019, 1019, 357, 0, 357, 357, 357, 1019, 1019, 357, 0, 357, 357, 357, 1019, 1019, 357, 0, 357, 357, 357, 1019, 1019, 357, 0, 357, 357, 357, 1019, 1019, 357], "M3F": [2462, 2462, 0, 862, 862, 862, 862, 2462, 2462, 0, 862, 862, 862, 862, 2462, 2462, 0, 862, 862, 862, 862, 2462, 2462, 0, 862, 862, 862, 862, 2462, 2462, 0], "AquaFoz": [2315, 2315, 810, 810, 810, 810, 810, 2315, 2315, 810, 810, 810, 810, 810, 2315, 2315, 810, 810, 810, 810, 810, 2315, 2315, 810, 810, 810, 810, 810, 2315, 2315, 810], "BioParque": [3816, 3816, 0, 0, 1335, 1335, 1335, 3816, 3816, 0, 0, 1335, 1335, 1335, 3816, 3816, 0, 0, 1335, 1335, 1335, 3816, 3816, 0, 0, 1335, 1335, 1335, 3816, 3816, 0], "AquaRio": [4678, 3236, 1951, 2355, 1970, 1884, 2474, 4678, 3236, 1951, 2355, 1970, 1884, 2474, 4678, 3236, 1951, 2355, 1970, 1884, 2474, 4678, 3236, 1951, 2355, 1970, 1884, 2474, 4678, 3236, 1951], "Vila Velha": [336, 336, 118, 0, 118, 118, 118, 336, 336, 118, 0, 118, 118, 118, 336, 336, 118, 0, 118, 118, 118, 336, 336, 118, 0, 118, 118, 118, 336, 336, 118], "PNI": [6816, 6544, 3308, 4438, 4544, 4625, 5721, 6816, 6544, 3308, 4438, 4544, 4625, 5721, 6816, 6544, 3308, 4438, 4544, 4625, 5721, 6816, 6544, 3308, 4438, 4544, 4625, 5721, 6816, 6544, 3308], "Paineiras": [3751, 4112, 3561, 2665, 3717, 3660, 3828, 3751, 4112, 3561, 2665, 3717, 3660, 3828, 3751, 4112, 3561, 2665, 3717, 3660, 3828, 3751, 4112, 3561, 2665, 3717, 3660, 3828, 3751, 4112, 3561]}, "SETEMBRO": {"Três Pescadores": [0, 376, 376, 376, 1074, 1074, 1074, 0, 376, 376, 376, 1074, 1074, 376, 0, 376, 376, 376, 1074, 1074, 376, 0, 376, 376, 376, 1074, 1074, 376, 0, 376], "M3F": [1102, 1102, 1102, 1102, 2561, 2561, 0, 1268, 1102, 1102, 1102, 2561, 2561, 0, 1102, 1102, 1102, 1102, 2561, 2561, 0, 1102, 1102, 1102, 1102, 2561, 2561, 0, 1102, 1102], "AquaFoz": [1021, 1021, 1021, 1021, 2219, 2219, 2423, 1021, 1021, 1021, 1021, 2219, 2219, 1021, 1021, 1021, 1021, 1021, 2219, 2219, 1021, 1021, 1021, 1021, 1021, 2219, 2219, 1021, 1021, 1021], "BioParque": [0, 1249, 1249, 1249, 3568, 3568, 0, 0, 1249, 1249, 1249, 3568, 3568, 0, 0, 1249, 1249, 1249, 3568, 3568, 0, 0, 1249, 1249, 1249, 3568, 3568, 0, 0, 1249], "AquaRio": [2101, 1993, 2422, 2473, 4021, 3574, 4127, 2101, 1993, 2422, 2473, 4021, 3574, 2127, 2101, 1993, 2422, 2473, 4021, 3574, 2127, 2101, 1993, 2422, 2473, 3021, 2574, 2127, 2101, 1993], "Vila Velha": [0, 147, 147, 147, 420, 420, 420, 0, 147, 147, 147, 420, 420, 147, 0, 147, 147, 147, 420, 420, 147, 0, 147, 147, 147, 420, 420, 147, 0, 147], "PNI": [4509, 5233, 5135, 5159, 7572, 7692, 6880, 4509, 5233, 5135, 5159, 7572, 7692, 3880, 4509, 5233, 5135, 5159, 7572, 7692, 3880, 4509, 4233, 4135, 5159, 7572, 7692, 3880, 4509, 4233], "Paineiras": [2892, 3301, 3120, 4542, 5000, 3360, 3998, 2892, 3301, 3120, 4542, 5000, 3360, 3998, 2892, 3301, 3120, 4542, 5000, 3360, 3998, 2892, 3301, 3120, 4542, 5000, 3360, 3998, 2892, 3301]}, "OUTUBRO": {"Três Pescadores": [322, 322, 920, 920, 322, 0, 322, 322, 322, 920, 920, 920, 0, 322, 322, 322, 920, 920, 322, 0, 322, 322, 322, 920, 920, 322, 0, 322, 322, 322, 920], "M3F": [1104, 1169, 2199, 2199, 0, 1104, 1104, 1104, 1104, 2199, 2199, 2199, 1104, 1104, 1104, 1104, 2199, 2199, 0, 1104, 1104, 1104, 1104, 2199, 2199, 0, 1104, 1104, 1104, 1104, 2199], "AquaFoz": [1115, 1115, 2493, 2493, 1115, 1115, 1115, 1115, 1115, 2493, 2493, 2493, 1115, 1115, 1115, 1115, 2493, 2493, 1115, 1115, 1115, 1115, 1115, 2493, 2493, 1115, 1115, 1115, 1115, 1115, 2484], "BioParque": [1501, 1501, 4729, 4729, 0, 0, 1501, 1501, 1501, 4729, 4729, 8000, 0, 1501, 1501, 1501, 4729, 4729, 0, 0, 1501, 1501, 1501, 4729, 4729, 0, 0, 1501, 1501, 1501, 4729], "AquaRio": [2431, 2742, 4725, 5569, 2503, 2600, 3349, 2431, 2742, 4725, 5569, 4503, 2600, 3349, 2431, 2742, 4725, 5569, 2503, 2600, 3349, 2431, 2742, 3725, 4569, 2503, 2600, 3349, 2431, 2742, 4725], "Vila Velha": [124, 124, 355, 355, 124, 0, 124, 124, 124, 355, 355, 355, 0, 124, 124, 124, 355, 355, 124, 0, 124, 124, 124, 355, 355, 124, 0, 124, 124, 124, 355], "PNI": [5108, 5838, 8611, 7465, 5486, 6140, 4799, 5108, 5838, 8611, 7465, 5486, 6140, 4799, 5108, 5838, 8611, 7465, 5486, 6140, 4799, 5108, 5838, 8611, 7465, 5486, 6140, 4799, 5108, 5838, 8611], "Paineiras": [3121, 3623, 4304, 4389, 4577, 3131, 2398, 3121, 3623, 4304, 4389, 4577, 3131, 2398, 3121, 3623, 4304, 4389, 4577, 3131, 2398, 3121, 3623, 4304, 4389, 4577, 3131, 2398, 3121, 3623, 4304]}, "NOVEMBRO": {"Três Pescadores": [1069, 1069, 0, 374, 374, 374, 1069, 1069, 374, 0, 374, 374, 374, 1069, 1069, 374, 0, 374, 1230, 1069, 1069, 1069, 374, 0, 374, 374, 374, 1069, 1069, 374], "M3F": [2444, 2443, 1254, 1254, 1254, 1254, 2443, 2443, 0, 1254, 1254, 1254, 1254, 2443, 2443, 0, 1254, 1254, 1254, 1254, 2443, 2443, 0, 1254, 1254, 1254, 1254, 2443, 2443, 0], "AquaFoz": [2665, 2665, 1101, 1101, 1101, 1101, 2665, 2665, 1101, 1101, 1101, 1101, 1101, 2665, 2665, 1101, 1101, 1101, 2271, 2665, 2665, 2665, 1101, 1101, 1101, 1101, 1101, 2665, 2665, 999], "BioParque": [2627, 2627, 0, 920, 920, 920, 2627, 2627, 0, 0, 920, 920, 920, 2627, 2627, 0, 0, 920, 3021, 2627, 2627, 2627, 0, 0, 920, 920, 920, 2627, 2627, 0], "AquaRio": [3402, 3548, 1998, 1755, 2046, 2732, 3126, 3002, 2648, 1998, 1755, 2046, 2732, 3126, 3002, 2648, 1998, 1755, 2046, 4232, 3126, 3002, 2648, 1998, 1755, 2046, 2732, 3126, 3002, 2648], "Vila Velha": [456, 456, 0, 160, 160, 160, 456, 456, 160, 0, 160, 160, 160, 456, 456, 160, 0, 160, 456, 524, 456, 456, 160, 0, 160, 160, 160, 456, 456, 160], "PNI": [9084, 5112, 5934, 5232, 6384, 7437, 8572, 9084, 5112, 5934, 5232, 6384, 7437, 8572, 9084, 5112, 5934, 5232, 6384, 7437, 8572, 9084, 5112, 5934, 5232, 6384, 7437, 8572, 9084, 5112], "Paineiras": [3391, 3695, 3327, 4177, 4074, 5648, 5323, 3391, 3695, 3327, 4177, 4074, 5648, 5323, 3391, 3695, 3327, 4177, 4074, 5648, 5323, 3391, 3695, 3327, 4177, 4074, 5648, 5323, 3391, 3695]}, "DEZEMBRO": {"Três Pescadores": [0, 393, 393, 393, 852, 852, 393, 0, 393, 393, 393, 852, 852, 393, 0, 393, 393, 393, 1124, 1124, 420, 420, 420, 1293, 1124, 1124, 1124, 470, 470, 435, 420], "M3F": [1121, 1121, 1121, 1121, 1950, 1950, 0, 1121, 1121, 1121, 1121, 1950, 1950, 0, 1121, 1121, 1121, 1121, 2010, 2010, 2010, 2010, 2010, 0, 2745, 2745, 2745, 2010, 2069, 2500, 0], "AquaFoz": [1042, 1042, 1042, 1042, 2521, 2521, 1042, 1042, 1042, 1042, 1042, 2521, 2521, 1042, 1042, 1042, 1042, 1042, 2977, 2977, 1042, 1042, 1042, 1042, 2000, 2977, 2977, 2400, 2400, 2400, 2151], "BioParque": [0, 1101, 1101, 1101, 2625, 2625, 0, 0, 1101, 1101, 1101, 2625, 2510, 0, 0, 1101, 1101, 1101, 2525, 2525, 1300, 1300, 1300, 0, 1900, 2625, 2625, 1650, 1650, 1650, 0], "AquaRio": [1852, 1862, 1905, 1943, 2621, 2281, 1852, 1862, 1905, 1943, 1943, 3140, 2281, 1943, 1852, 1862, 1905, 1943, 3251, 3096, 1852, 1862, 1905, 1351, 2251, 3685, 3950, 3482, 3485, 3489, 2051], "Vila Velha": [0, 148, 148, 148, 423, 423, 148, 0, 148, 148, 148, 423, 423, 148, 0, 148, 148, 148, 423, 423, 148, 0, 148, 148, 246, 423, 423, 240, 240, 240, 240], "PNI": [5661, 6053, 5842, 5776, 7084, 7008, 5601, 5661, 6053, 5842, 5776, 7984, 7908, 5601, 5661, 6053, 5842, 5776, 8984, 8908, 5661, 6053, 5842, 7893, 8280, 8280, 8084, 7680, 7540, 6852, 7152], "Paineiras": [3004, 3483, 3471, 4076, 3625, 3690, 3203, 3004, 3983, 4471, 4476, 3625, 3690, 3203, 3504, 3983, 4471, 4476, 3625, 3690, 3803, 3404, 3983, 2671, 4476, 3425, 3490, 4203, 4504, 4983, 4471]}}''')


def build_visitacao_meta(service, spreadsheet_id, sheet_name, meses_meta):
    """Le a tabela dinamica 'Soma de Nova2026' (uma coluna por dia do ano, uma linha por
    parque) e devolve {mes: {parque: [valor_dia1, valor_dia2, ...]}} só para os meses
    informados (tipicamente os que ainda nao tem resultado real)."""
    if not meses_meta:
        return {}
    rows = get_values(service, spreadsheet_id, sheet_name)

    # Acha a tabela dinamica de verdade: a aba tem MAIS DE UMA celula solta chamada
    # "Soma de Nova2026" (ex.: uma mini-tabela de referencia perto do topo, com "Rótulos
    # de Linha" tambem na linha seguinte, mas SEM colunas de data -- so' nomes de parque).
    # Por isso nao paramos no primeiro match: só aceitamos o candidato cuja linha de
    # cabecalho realmente tem rótulos de data (formato "01/mar") nas colunas seguintes.
    col_to_date = {}
    header_row_idx, label_col = None, None
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            if val != "Soma de Nova2026":
                continue
            next_row = rows[i + 1] if i + 1 < len(rows) else []
            if cell(next_row, j) != "Rótulos de Linha":
                continue
            candidato = {}
            for c in range(j + 1, len(next_row)):
                dt = _parse_dd_mon(cell(next_row, c))
                if dt is not None:
                    candidato[c] = dt
            if candidato:
                header_row_idx, label_col, col_to_date = i + 1, j, candidato
                break
        if header_row_idx is not None:
            break
    if header_row_idx is None:
        return {}

    result = {mes: {} for mes in meses_meta}
    for mes in meses_meta:
        n_days = calendar.monthrange(2026, MONTH_NUMBER[mes])[1]
        for park in set(VISITACAO_META_PARK_NAMES.values()):
            result[mes][park] = [None] * n_days

    for r in range(header_row_idx + 1, len(rows)):
        label = cell(rows[r], label_col)
        if not isinstance(label, str):
            continue
        if label.strip().lower().startswith("total"):
            break
        park = VISITACAO_META_PARK_NAMES.get(label)
        if park is None:
            continue
        for c, dt in col_to_date.items():
            mes_en = next((m for m in meses_meta if MONTH_NUMBER[m] == dt.month), None)
            if mes_en is None:
                continue
            v = cell(rows[r], c)
            if isinstance(v, (int, float)):
                result[mes_en][park][dt.day - 1] = v
    return result


# ---------------------------------------------------------------------------
# Visitação Parques 2026.xlsx -> aba "Mapa Clima"
# Serie diaria (sequencial, ano inteiro) com o codigo de clima do dia, por regiao. O
# painel so' usa 1 "parque representante" por regiao (CLIMA_REGIOES no HTML): AquaRio
# representa Parques Rio, PNI representa Parques Foz, Vila Velha e Três Pescadores tem
# regiao propria. BUG EVITADO: essa aba nunca foi ligada ao pipeline -- por isso a aba
# Clima do painel sumia (ficava sem nenhum dado) assim que o painel passava a ler o
# data.json em vez dos numeros fixos.
# ---------------------------------------------------------------------------

# (coluna da Data, coluna do texto "Clima") por parque-representante, 0-indexed
CLIMA_REGIAO_COLS = {
    "AquaRio": (8, 9),
    "PNI": (17, 18),
    "Vila Velha": (26, 27),
    "Três Pescadores": (32, 33),
}

CLIMA_LABEL_TO_CODE = {
    "dia de sol": "sun", "nublado": "cloud", "nublado com chuva": "cloud-rain",
    "chuvoso": "rain", "sol + chuva": "sun-rain", "tempestade": "storm",
    "calor excessivo": "heat", "frio intenso": "cold", "sol + vento": "wind-sun",
    "chuva + vento": "wind-rain", "ventoso": "wind", "fechado": "closed", "alerta": "alert",
    # BUG EVITADO (10/08/2026): "Sol + Nuvem" é um rótulo usado na planilha que não tinha
    # entrada aqui -- por isso os dias com esse texto (ex.: AquaRio 02-04/08, PNI 02/08)
    # ficavam sem código nenhum, e a aba Clima do painel mostrava esses dias como "sem
    # dado" mesmo com a planilha preenchida. Novo código "sun-cloud" (ver CODE_LABEL/
    # CODE_EMOJI/CLIMA_COLORS no index.html).
    "sol + nuvem": "sun-cloud",
    # BUG EVITADO (11/08/2026): a fórmula da planilha passou a gerar 3 novos rótulos
    # combinados com "+ Alerta" que não tinham entrada aqui -- por isso dias como
    # AquaRio 06-07/08/2026 ("Dia de Sol + Alerta" e "Nublado + Alerta") e AquaRio
    # 29/07/2026 ("Ventoso + Alerta") ficavam sem código, e a aba Clima mostrava esses
    # dias como "sem dado". Novos códigos "sun-alert"/"cloud-alert"/"wind-alert" (ver
    # CODE_LABEL/CODE_EMOJI/CLIMA_COLORS no index.html).
    "ventoso + alerta": "wind-alert",
    "dia de sol + alerta": "sun-alert",
    "nublado + alerta": "cloud-alert",
}


def build_clima_emoji(service, spreadsheet_id, sheet_name, meses_com_dados):
    rows = get_values(service, spreadsheet_id, sheet_name)
    result = {mes: {} for mes in meses_com_dados}
    for mes in meses_com_dados:
        month_number = MONTH_NUMBER[mes]
        n_days = calendar.monthrange(2026, month_number)[1]
        for park in CLIMA_REGIAO_COLS:
            result[mes][park] = [None] * n_days

    for row in rows:
        for park, (date_col, clima_col) in CLIMA_REGIAO_COLS.items():
            date_val = cell(row, date_col)
            dt = None
            if isinstance(date_val, datetime.datetime):
                dt = date_val.date()
            elif isinstance(date_val, (int, float)):
                dt = serial_to_date(date_val)
            if dt is None or dt.year != 2026:
                continue
            mes_en = next((m for m in meses_com_dados if MONTH_NUMBER[m] == dt.month), None)
            if mes_en is None:
                continue
            label = cell(row, clima_col)
            if not isinstance(label, str) or not label.strip():
                continue
            code = CLIMA_LABEL_TO_CODE.get(label.strip().lower())
            if code:
                result[mes_en][park][dt.day - 1] = code
    return result


# ---------------------------------------------------------------------------
# CAPTAÇÃO CV - 3P (mesma planilha "Visitação Parques 2026")
# ---------------------------------------------------------------------------

def build_captacao_cv_3p(service, spreadsheet_id, sheet_name):
    """Le a aba "CAPTAÇÃO CV - 3P": um bloco de 4 colunas (DATA | VISITAÇÃO 3P | CV |
    CAPTAÇÃO) por mes, um do lado do outro, com o titulo do mes uma linha acima do
    cabecalho (ex.: "CAPTAÇÃO TRÊS PESCADORES AGOSTO").

    BUG EVITADO: a versao anterior usava um dict fixo de indices de coluna por mes
    (CAPTACAO_CV_3P_COLS), que so' ia até Julho -- todo mes novo (Agosto em diante)
    tinha que ser adicionado a mao nesse dict, e enquanto isso o card ficava mostrando
    "-" (mes "sem dado", mesmo com a planilha ja preenchida). Agora o bloco de cada
    mes e' encontrado dinamicamente pelo titulo (mesmo padrao usado em outras partes
    do pipeline, ex.: build_evolucao_ppt_parcial), sem precisar editar nada todo mes.

    Alem do total por mes (by_month) e do acumulado (anual), tambem devolve o detalhe
    DIARIO (diario: {"AAAA-MM-DD": {"visitacao":.., "cv":..}}) -- usado pelo frontend
    pra somar qualquer intervalo de datas (filtro "Personalizado"), que antes so'
    funcionava pra Visitação/OBZ e sempre mostrava "-" na Captação CV.
    """
    rows = get_values(service, spreadsheet_id, sheet_name)
    if len(rows) < 2:
        return {}, {"visitacao": 0, "cv": 0}, {}

    titulo_row, header_row = rows[0], rows[1]
    meses_norm = {_norm_txt(nome): (nome, num) for nome, num in MONTH_NUMBER.items()}

    blocos = []  # (mes_nome, month_number, dcol)
    for c, v in enumerate(header_row):
        if _norm_txt(v) != "DATA":
            continue
        titulo = _norm_txt(cell(titulo_row, c))
        for nome_norm, (mes_nome, month_number) in meses_norm.items():
            if nome_norm and nome_norm in titulo:
                blocos.append((mes_nome, month_number, c))
                break

    by_month = {}
    diario = {}
    total_vis, total_cv = 0.0, 0.0
    for mes_nome, month_number, dcol in blocos:
        vcol, ccol = dcol + 1, dcol + 2
        vis_sum, cv_sum = 0.0, 0.0
        for r in rows[2:]:
            d = serial_to_date(cell(r, dcol))
            # BUG EVITADO: filtramos por DATA real do mes, nao por posicao de linha --
            # a planilha tem uma linha de "total do mes" logo apos os dias, que nao tem
            # data preenchida. Se somarmos por posicao de linha esse total entra junto
            # e o resultado sai em dobro.
            if d and d.year == 2026 and d.month == month_number:
                v = cell(r, vcol)
                c = cell(r, ccol)
                if isinstance(v, (int, float)):
                    vis_sum += v
                if isinstance(c, (int, float)):
                    cv_sum += c
                diario[d.isoformat()] = {
                    "visitacao": v if isinstance(v, (int, float)) else None,
                    "cv": c if isinstance(c, (int, float)) else None,
                }
        by_month[mes_nome] = {"visitacao": int(vis_sum), "cv": int(cv_sum)}
        total_vis += vis_sum
        total_cv += cv_sum
    return by_month, {"visitacao": int(total_vis), "cv": int(total_cv)}, diario


# ---------------------------------------------------------------------------
# Share E-commerce_Parques Rio - _2026.xlsx -> aba "Share_Ecommerce_2026"
# Serie mensal historica (Jan/2023 em diante) por parque: Visitação total, Ecommerce,
# Share, R$ em mídia. Uma coluna por mes/ano. E' a fonte de "investimentoMidia.meses".
# ---------------------------------------------------------------------------

# linha (0-indexed) onde comeca o bloco de cada parque nesta aba. "3P" (Três Pescadores)
# e "Vila Velha" ficam num grupo separado, abaixo do titulo "SOUL PARQUES" -- so' tem
# coluna 2026 preenchida (parques novos, sem historico 2025), mas o layout de linhas
# (Visitação/Ecommerce/Share/R$ em mídia) e' o mesmo dos outros blocos. As chaves "3P" e
# "Vila Velha" (em vez de "Três Pescadores") sao de proposito -- e' o nome que o HTML
# (INVEST_PARKS) ja espera pra esses dois parques nesta tabela especifica.
# BUG EVITADO (mesma classe do fix em parse_month_summary): antes, os blocos eram achados
# por OFFSET FIXO (SHARE_ECOMMERCE_BLOCKS abaixo, mantido só como fallback/documentação).
# Isso quebra se algum bloco for removido/inserido na planilha -- ex.: BioParque saiu do
# Grupo Cataratas e o time pode remover o bloco dele aqui também, empurrando AquaRio/
# Paineiras/M3F/etc. pra cima, o que faria cada parque ler o resultado do PRÓXIMO. Agora
# cada bloco é achado pelo próprio rótulo (coluna A), via _find_share_ecommerce_blocks().
SHARE_ECOMMERCE_BLOCKS = {
    "BioParque": 0, "AquaRio": 9, "Paineiras": 17, "M3F": 25, "AquaFoz": 33,
    "3P": 44, "Vila Velha": 52,
}
MESES_PT = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho",
            "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]


def _find_share_ecommerce_blocks(rows):
    """Acha a linha de título (r0) de cada bloco de parque na aba Share_Ecommerce_2026
    pelo próprio rótulo da coluna A, em vez de confiar em offsets fixos -- funciona com
    qualquer bloco removido/inserido/reordenado (ver comentário acima de
    SHARE_ECOMMERCE_BLOCKS)."""
    labels_to_park = {
        "BioParque": "BioParque", "AquaRio": "AquaRio", "Paineiras": "Paineiras",
        "M3F": "M3F", "AquaFoz": "AquaFoz", "3P": "3P", "Três Pescadores": "3P",
        "Vila Velha": "Vila Velha",
    }
    blocos = {}
    for i, row in enumerate(rows):
        label = cell(row, 0)
        if isinstance(label, str) and label.strip() in labels_to_park:
            park = labels_to_park[label.strip()]
            if park not in blocos:
                blocos[park] = i
    return blocos


def _find_labeled_row_idx(rows, r0, label, max_offset=10):
    """Acha o indice (em `rows`) da linha rotulada `label` na coluna A, procurando a
    partir de r0+1 ate r0+max_offset. Usado pra achar "Ecommerce (base TI)" cujo
    deslocamento varia por bloco (BioParque tem uma linha extra "Historico Pareto" que
    empurra o resto do bloco pra baixo, os outros parques nao tem)."""
    label_norm = label.strip().lower()
    limit = min(r0 + max_offset, len(rows) - 1)
    for idx in range(r0 + 1, limit + 1):
        val = cell(rows[idx], 0)
        if isinstance(val, str) and val.strip().lower() == label_norm:
            return idx
    return None


def _norm_txt(s):
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return s.strip().upper()


EVOLUCAO_PPT_PARK_ALIASES = {
    "AQUARIO": "AquaRio",
    "BIOPARQUE": "BioParque",
    "PAINEIRAS": "Paineiras",
    "AQUAFOZ": "AquaFoz",
}


def build_evolucao_ppt_parcial(service, spreadsheet_id, mes_atual_pt_upper, ano_2025=2025, cutoff_day=None):
    """Comparativo "Vs 2025" proporcional (D-1) pro mês vigente.

    A linha "Ecommerce (base TI)" (usada pros meses já fechados) só tem o total do mês
    inteiro -- não dá pra comparar "mesmo período" com ela enquanto o mês em curso (2026)
    ainda não terminou. Por pedido do usuário, pro mês vigente a gente usa em vez disso a
    aba "EVOLUÇÃO<MÊS> (PPT)" (criada manualmente todo mês, mesmo padrão de nome -- ex.:
    "EVOLUÇÃOAGOSTO (PPT)" a partir de 01/08/2026), que tem o detalhe DIÁRIO de 2025, e soma
    Visitação/Ecommerce de 2025 do dia 1 até D-1 (o mesmo período que o 2026 parcial já
    cobre). Só é usado pro mês vigente -- os meses fechados continuam com a Ecommerce (base
    TI) mensal cheia, sem mudança.

    O layout da aba tem um bloco de 4 colunas por parque (DIA | VISITAÇÃO 2025 |
    E-COMMERCE 2025 | SHARE), procurado dinamicamente pelo cabeçalho "DIA" -- o rótulo do
    parque (uma linha acima) varia de nome por parque (ex.: o bloco do M3F aparece rotulado
    "MARCO DAS TRÊS FRONTEIRAS"), por isso o match é por trecho de texto, não código exato.
    "3P"/"Vila Velha" não tem esse bloco (parques novos, sem 2025) -- ficam de fora, sem erro.
    """
    if cutoff_day is None:
        hoje_brt = (datetime.datetime.utcnow() - datetime.timedelta(hours=3)).date()
        cutoff_day = hoje_brt.day - 1
    if cutoff_day < 1:
        return {}

    mes_num = MONTH_NUMBER.get(mes_atual_pt_upper)
    if mes_num is None:
        return {}

    sheet_name = f"EVOLUÇÃO{mes_atual_pt_upper} (PPT)"
    rows = get_values(service, spreadsheet_id, sheet_name)
    if not rows:
        return {}

    header_row = None
    for r, row in enumerate(rows):
        if any(_norm_txt(cell(row, c)) == "DIA" for c in range(len(row))):
            header_row = r
            break
    if header_row is None:
        return {}

    header = rows[header_row]
    titulo_row = rows[header_row - 1] if header_row > 0 else []
    blocos = {}
    for c, v in enumerate(header):
        if _norm_txt(v) != "DIA":
            continue
        v1 = _norm_txt(cell(header, c + 1))
        v2 = _norm_txt(cell(header, c + 2)).replace("-", "")
        if "VISITA" not in v1 or "ECOMMERCE" not in v2:
            continue
        titulo = _norm_txt(cell(titulo_row, c))
        park = None
        for alias, canon in EVOLUCAO_PPT_PARK_ALIASES.items():
            if alias in titulo:
                park = canon
                break
        if park is None and "MARCO" in titulo and "FRONTEIRA" in titulo:
            park = "M3F"
        if park:
            blocos[park] = c

    # Retorna os totais BRUTOS (parcial D-1 e mês inteiro) direto da aba Evolução, sem
    # nenhuma correção de escala ainda -- essa aba é uma fonte separada da "Ecommerce (base
    # TI)" e pode não fechar exatamente no mesmo total do mês (metodologias diferentes).
    # Quem usa isso (build_investimento_midia / build_evolucao_mensal) escala o parcial pela
    # razão entre o total oficial (Ecommerce base TI, que é o número que o resto do painel já
    # trata como correto) e esse total bruto do mês inteiro aqui -- ver _escala_proporcional.
    resultado = {}
    for park, dia_col in blocos.items():
        total_ecom_parcial, total_vis_parcial = 0.0, 0.0
        total_ecom_total, total_vis_total = 0.0, 0.0
        achou = False
        for r in range(header_row + 1, len(rows)):
            serial = cell(rows[r], dia_col)
            if not isinstance(serial, (int, float)):
                continue
            d = serial_to_date(serial)
            if d is None or d.year != ano_2025 or d.month != mes_num:
                continue
            ecv = cell(rows[r], dia_col + 2)
            visv = cell(rows[r], dia_col + 1)
            if isinstance(ecv, (int, float)):
                total_ecom_total += ecv
                if d.day <= cutoff_day:
                    total_ecom_parcial += ecv
                achou = True
            if isinstance(visv, (int, float)):
                total_vis_total += visv
                if d.day <= cutoff_day:
                    total_vis_parcial += visv
        # Guarda de segurança: se o bloco de 2025 desse parque na aba PPT tiver o mês
        # inteiro fechado mas o trecho até o cutoff_day vier zerado (ex.: alguem colou o
        # 2025 inteiro de uma vez só perto do fim do mes, em vez de dia a dia -- caso real
        # visto no BioParque em Agosto/2026, onde os dias 1-20 vieram 0 e só 21-31 tinham
        # valor), a "regra de 3" geraria um comparativo falso de 0 em vez de cair pro total
        # do mes inteiro (que já está correto). Por isso só usamos o comparativo D-1 quando
        # o parcial em si é > 0 -- senão, melhor deixar o override de fora e manter o valor
        # cheio de "Ecommerce (base TI)" que o resto do painel já usa.
        if achou and total_vis_total > 0 and total_vis_parcial > 0:
            resultado[park] = {
                "visitacao2025_parcial": total_vis_parcial,
                "ecommerce2025_parcial": total_ecom_parcial,
                "visitacao2025_total": total_vis_total,
                "ecommerce2025_total": total_ecom_total,
            }
    return resultado


def _escala_proporcional(parcial, bruto_total_evolucao, oficial_base_ti):
    """Regra de 3 (pedido do usuário): corrige o parcial D-1 da aba Evolução pela razão entre
    o total "oficial" (Ecommerce base TI, mês fechado) e o total bruto do mês inteiro
    conforme a própria aba Evolução -- evita que o comparativo do mês vigente divirja do
    número que o resto do painel já usa como correto pra esse mesmo mês quando ele fechar.
    Sem essa correção, um parque com metodologias diferentes entre as duas abas mostraria
    Jul/25 subindo dia a dia até um valor, e depois "pulando" pro número oficial quando o
    mês fecha de verdade -- com a correção, os dois já ficam na mesma escala desde já.
    Sem ancora (base TI vazio) ou sem total bruto (mês sem dado na aba Evolução), devolve o
    parcial puro, sem inventar número."""
    if not bruto_total_evolucao or oficial_base_ti is None:
        return parcial
    return parcial * (oficial_base_ti / bruto_total_evolucao)


def build_investimento_midia(service, spreadsheet_id, sheet_name, meses_com_dados, ppt_parcial=None):
    """Monta SHARE.investimentoMidia.meses para os parques rastreados nesta aba
    (AquaRio, BioParque, Paineiras, M3F, AquaFoz, e os dois do bloco "SOUL PARQUES":
    "3P"/Três Pescadores e Vila Velha).

    Para meses fechados (todo mes exceto o corrente), tanto o lado 2025 quanto o 2026 sao
    o total do mes inteiro. Para o mes corrente (ainda em andamento), o lado 2026 e'
    parcial (só os dias já lançados na planilha) e o 2025 e' o mes inteiro do ano passado
    -- usado como referencia de "para onde estamos indo", nao como comparação dia-a-dia.

    "3P" e "Vila Velha" so' tem coluna 2026 (parques novos, sem historico 2025) -- os
    campos *2025 ficam None/0 automaticamente pra eles, sem precisar de tratamento
    especial (num() ja devolve None pra celula vazia).
    """
    rows = get_values(service, spreadsheet_id, sheet_name)
    blocos = _find_share_ecommerce_blocks(rows) or SHARE_ECOMMERCE_BLOCKS
    meses = {}
    for i, mes_en in enumerate(meses_com_dados):
        mes_idx = MONTH_NUMBER[mes_en] - 1  # 0 = Janeiro
        mes_pt = MESES_PT[mes_idx]
        idx_2026 = 37 + mes_idx  # coluna do mes/ano em 2026 (Jan/2026 comeca no indice 37)
        idx_2025 = 25 + mes_idx  # coluna do mes/ano em 2025 (Jan/2025 comeca no indice 25)
        meses[mes_pt] = {}
        for park, r0 in blocos.items():
            if park == "BioParque" and not _bioparque_ainda_conta(2026, mes_idx + 1):
                continue  # saiu do Grupo Cataratas -- some do Share E-commerce a partir daqui
            vis_row, inv_row = rows[r0 + 1], rows[r0 + 4]
            # BUG EVITADO: por pedido do usuario, o painel passou a usar "Ecommerce (base
            # TI)" em vez da linha "Ecommerce" simples -- mas o deslocamento dessa linha
            # varia por bloco (BioParque tem uma linha extra "Historico Investimento em
            # midia (Pareto)" que empurra o resto do bloco pra baixo), entao procuramos
            # pelo rotulo em vez de usar um indice fixo.
            ecom_ti_idx = _find_labeled_row_idx(rows, r0, "Ecommerce (base TI)")
            ecom_row = rows[ecom_ti_idx] if ecom_ti_idx is not None else rows[r0 + 2]
            share_row = rows[ecom_ti_idx + 1] if ecom_ti_idx is not None else rows[r0 + 3]

            def num(row, idx):
                v = cell(row, idx)
                return v if isinstance(v, (int, float)) else None

            vis26, ecom26 = num(vis_row, idx_2026), num(ecom_row, idx_2026)
            vis25, ecom25 = num(vis_row, idx_2025), num(ecom_row, idx_2025)
            meses[mes_pt][park] = {
                "visitacao2026": vis26,
                "ecommerce2026": ecom26,
                "share2026": (ecom26 / vis26) if (vis26 and ecom26 is not None) else num(share_row, idx_2026),
                "visitacao2025": vis25,
                "ecommerce2025": ecom25,
                "share2025": (ecom25 / vis25) if (vis25 and ecom25 is not None) else num(share_row, idx_2025),
                "investimento2026": num(inv_row, idx_2026),
                "investimento2025": num(inv_row, idx_2025),
            }
            # Mês vigente: troca o lado 2025 pelo comparativo proporcional (D-1) da aba
            # "EVOLUÇÃO<MÊS> (PPT)", já escalado (regra de 3) pro mesmo total oficial de
            # Ecommerce (base TI) -- ver build_evolucao_ppt_parcial / _escala_proporcional.
            if mes_en == meses_com_dados[-1] and ppt_parcial and park in ppt_parcial:
                p = ppt_parcial[park]
                vis25_corrigido = _escala_proporcional(p["visitacao2025_parcial"], p["visitacao2025_total"], vis25)
                ecom25_corrigido = _escala_proporcional(p["ecommerce2025_parcial"], p["ecommerce2025_total"], ecom25)
                meses[mes_pt][park]["visitacao2025"] = vis25_corrigido
                meses[mes_pt][park]["ecommerce2025"] = ecom25_corrigido
                meses[mes_pt][park]["share2025"] = (ecom25_corrigido / vis25_corrigido) if vis25_corrigido else None
    return meses


def build_evolucao_mensal(service, spreadsheet_id, sheet_name, meses_com_dados, ppt_parcial=None, ano_inicio=2025, mes_inicio=1, ano_fim=2026):
    """Serie historica mes a mes (investimento, share, visitacaoTotal) de Jan/2025 ate o
    mes/ano atual, mesma aba "Share_Ecommerce_2026" — e' o mesmo dado de
    build_investimento_midia, só que olhando pra tras (nao comparando 2026 vs 2025 lado a
    lado, e sim uma linha do tempo unica).

    mes_fim agora vem de meses_com_dados (mês vigente) em vez de fixo -- antes ficava
    hardcoded como "7" (Julho) e precisava ser editado a mão todo mês."""
    mes_fim = MONTH_NUMBER[meses_com_dados[-1]] if meses_com_dados else 7
    rows = get_values(service, spreadsheet_id, sheet_name)
    labels = []
    idx_por_label = []
    y, m = ano_inicio, mes_inicio
    while (y, m) <= (ano_fim, mes_fim):
        labels.append(f"{m:02d}/{y % 100:02d}")
        idx_por_label.append(25 + (y - 2025) * 12 + (m - 1))
        m += 1
        if m > 12:
            m = 1
            y += 1

    label_alvo = f"{mes_fim:02d}/{ano_inicio % 100:02d}"
    idx_alvo = labels.index(label_alvo) if label_alvo in labels else None

    def num(row, idx):
        v = cell(row, idx)
        return v if isinstance(v, (int, float)) else None

    blocos = _find_share_ecommerce_blocks(rows) or SHARE_ECOMMERCE_BLOCKS
    parques = {}
    for park, r0 in blocos.items():
        vis_row, inv_row = rows[r0 + 1], rows[r0 + 4]
        # mesma logica de build_investimento_midia: usa a Share de "Ecommerce (base TI)",
        # cujo deslocamento varia por bloco -- procura pelo rotulo em vez de indice fixo.
        ecom_ti_idx = _find_labeled_row_idx(rows, r0, "Ecommerce (base TI)")
        ecom_row = rows[ecom_ti_idx] if ecom_ti_idx is not None else rows[r0 + 2]
        share_row = rows[ecom_ti_idx + 1] if ecom_ti_idx is not None else rows[r0 + 3]

        parques[park] = {
            "investimento": [num(inv_row, i) for i in idx_por_label],
            "share": [num(share_row, i) for i in idx_por_label],
            "visitacaoTotal": [num(vis_row, i) for i in idx_por_label],
        }

        # Mesmo ajuste de build_investimento_midia: no mês vigente, troca a entrada do
        # "mesmo mês do ano passado" (ex.: Julho/25) pelo comparativo proporcional D-1,
        # escalado (regra de 3) pro mesmo total oficial de Ecommerce (base TI).
        if ppt_parcial and park in ppt_parcial and idx_alvo is not None:
            p = ppt_parcial[park]
            col_alvo = idx_por_label[idx_alvo]
            vis_base_ti = num(vis_row, col_alvo)
            ecom_base_ti = num(ecom_row, col_alvo)
            vis25_corrigido = _escala_proporcional(p["visitacao2025_parcial"], p["visitacao2025_total"], vis_base_ti)
            ecom25_corrigido = _escala_proporcional(p["ecommerce2025_parcial"], p["ecommerce2025_total"], ecom_base_ti)
            parques[park]["visitacaoTotal"][idx_alvo] = vis25_corrigido
            parques[park]["share"][idx_alvo] = (ecom25_corrigido / vis25_corrigido) if vis25_corrigido else None

        if park == "BioParque":
            # saiu do Grupo Cataratas a partir de Agosto/2026 -- zera os pontos da série a
            # partir daí (Jan/25-Jul/26 continuam intactos, é literalmente o histórico real).
            for i, label in enumerate(labels):
                mm, yy = label.split("/")
                ano_label = 2000 + int(yy)
                if not _bioparque_ainda_conta(ano_label, int(mm)):
                    parques[park]["investimento"][i] = None
                    parques[park]["share"][i] = None
                    parques[park]["visitacaoTotal"][i] = None

    return {"labels": labels, "parques": parques}


def build_share_meta_grupo_cataratas(investimento_midia_meses, meses_com_dados):
    """SHARE_META_GRUPO_CATARATAS: nao precisa de nenhuma planilha nova -- e' a soma dos 5
    parques com e-commerce (AquaRio, BioParque, Paineiras, M3F, AquaFoz) que ja lemos em
    build_investimento_midia. Antes esse bloco vivia hard-coded no HTML; agora e' calculado.

    OBSERVAÇÃO: comparando com os números fixos que estavam no HTML, Julho bate exatamente,
    mas alguns meses mais antigos (ex.: Junho) batem diferente -- a AquaFoz aparenta ter
    passado a ser rastreada no e-commerce só a partir de um certo mês, e o valor antigo
    hard-coded parece ter sido somado sem a AquaFoz nesses meses. Como não dá pra saber com
    certeza, a partir de agora, a soma é sempre com os 5 parques (mais transparente e
    consistente pra frente) -- isso pode mudar levemente os meses fechados mais antigos do
    gráfico de meta, mas o mês mais recente (o que importa pra acompanhar o dia a dia)
    sempre bate."""
    parks = ["AquaRio", "BioParque", "Paineiras", "M3F", "AquaFoz"]
    visitacao, ecommerce, share = [], [], []
    for mes_en in meses_com_dados:
        mes_pt = MESES_PT[MONTH_NUMBER[mes_en] - 1]
        d = investimento_midia_meses.get(mes_pt, {})
        vis_total = sum((d.get(p, {}).get("visitacao2026") or 0) for p in parks)
        ecom_total = sum((d.get(p, {}).get("ecommerce2026") or 0) for p in parks)
        visitacao.append(vis_total)
        ecommerce.append(ecom_total)
        share.append(ecom_total / vis_total if vis_total else None)
    return {"visitacao": visitacao, "ecommerce": ecommerce, "share": share}


# ---------------------------------------------------------------------------
# [2026] Mix OBZ e visitação.xlsx -> aba "ANÁLISE MIX DE ORIGEM 2026"
# Bloco "MIX PARCIAL | MES ATUAL": um mini-bloco por parque, com o nome do parque na
# coluna C e um cabecalho (2025/2026/Delta/Share) na mesma linha, seguido das linhas de
# categoria (Local/Brasileiros/Estrangeiros/... e "Total Mes" por ultimo). O nome do
# parque nesta aba as vezes difere do nome usado no resto do painel (ex.: "Aquario" sem
# acento/maiuscula) -- por isso o mapeamento MIX_ORIGEM_PARK_NAMES abaixo.
# BUG EVITADO (mesma classe dos outros): antes desta funcao, essa tabela inteira do
# painel (MIX_ORIGEM) nunca foi ligada ao pipeline -- ficava com numero fixo, travado
# na data em que foi colado a mao pela ultima vez, mesmo com o resto ja automatizado.
# ---------------------------------------------------------------------------

MIX_ORIGEM_PARK_NAMES = {
    "Aquario": "AquaRio",
    "BioParque": "BioParque",
    "Paineiras": "Paineiras",
    "M3F": "M3F",
    "AquaFoz": "AquaFoz",
    "Três Pescadores": "Três Pescadores",
    "Vila Velha": "Vila Velha",
    "PNI": "PNI",
}


def build_mix_origem(service, spreadsheet_id, sheet_name, col_offset=0):
    """col_offset=0 le' o bloco "MIX PARCIAL" (mes atual); col_offset=11 le' o bloco
    "MIX ACUMULADO ANO" que fica 11 colunas a direita, mesmas linhas, mesmo layout --
    usado pelo toggle Mensal/Acumulado do painel."""
    name_col = 2 + col_offset
    rows = get_values(service, spreadsheet_id, sheet_name)
    result = {}
    for i, row in enumerate(rows):
        nome_sheet = cell(row, name_col)
        if nome_sheet not in MIX_ORIGEM_PARK_NAMES:
            continue
        header_d = cell(row, name_col + 1)
        # confirma que e' mesmo a linha de cabecalho do bloco (coluna seguinte = 2025 ou
        # 2026), nao outra celula qualquer da planilha que por acaso tenha o mesmo nome
        if header_d not in (2025, 2026):
            continue
        formato = "full" if header_d == 2025 else "somente2026"
        park = MIX_ORIGEM_PARK_NAMES[nome_sheet]

        categorias = []
        j = i + 1
        while j < len(rows):
            label = cell(rows[j], name_col)
            if not label:
                break
            # BUG EVITADO: a planilha tem "Local " com espaco sobrando no final da celula;
            # sem o strip(), a cor da legenda (que faz lookup por nome exato) nao batia e
            # a barra/legenda de "Local" caia no cinza padrao.
            label = label.strip() if isinstance(label, str) else label
            if formato == "full":
                v2025, v2026 = cell(rows[j], name_col + 1), cell(rows[j], name_col + 2)
                delta = cell(rows[j], name_col + 3)
                share25, share26 = cell(rows[j], name_col + 4), cell(rows[j], name_col + 5)
            else:
                v2025, delta, share25 = None, None, None
                v2026, share26 = cell(rows[j], name_col + 1), cell(rows[j], name_col + 2)
            categorias.append({
                "label": label, "v2025": v2025, "v2026": v2026, "delta": delta,
                "share25": share25, "share26": share26,
            })
            is_total = isinstance(label, str) and label.strip().startswith("Total")
            j += 1
            if is_total:
                break
        result[park] = {"formato": formato, "categorias": categorias}
    return result


# ---------------------------------------------------------------------------
# [2026] Mix OBZ e visitação.xlsx -> abas "DIÁRIO - <PARQUE>"
# Ledger diário por parque (uma linha por dia do ano), com blocos de colunas repetidos
# por ano (2026 real, 2025 real, OBZ, anos anteriores, chuva etc.). Usamos só os 2
# primeiros blocos (2026 e 2025, sempre nessa ordem) pra alimentar o filtro por período
# customizado da aba Mix de Origem -- a aba "ANÁLISE MIX DE ORIGEM 2026" só tem o mês
# atual + acumulado do ano prontos, não tem cada mês passado em separado.
# BUG EVITADO (mesma classe de outras abas irregulares do projeto): o texto do cabeçalho
# da 2ª linha (ex.: "BRASIL 25") às vezes fica com o ano errado mesmo dentro do bloco de
# 2026 (a planilha não atualiza o sufixo do rótulo todo ano) -- por isso identificamos o
# bloco pelo ANO NA LINHA 1 (que está certo), nunca pelo sufixo do rótulo da linha 2.
# Categorias variam por parque (fiéis à fonte, sem forçar padronização -- mesmos nomes
# já usados na aba Mix de Origem mensal, pra herdar cor/legenda):
#   - AquaRio/BioParque/Paineiras/Três Pescadores/Vila Velha: Local, Brasileiros, Estrangeiros
#   - M3F/AquaFoz: Morador, Estado (coluna "LOCAL"), Brasileiros, Mercosul, Estrangeiros
#   - PNI: Morador/Estado (coluna única "MORADOR/LOCAL"), Brasileiros, Mercosul, Estrangeiros
# ---------------------------------------------------------------------------

DIARIO_SHEET_NAMES = {
    "AquaRio": "DIÁRIO - AQUA",
    "BioParque": "DIÁRIO - BIO",
    "Paineiras": "DIÁRIO - PAI",
    "M3F": "DIÁRIO - M3F",
    "AquaFoz": "DIÁRIO - AQF",
    "PNI": "DIÁRIO PNI",
    "Três Pescadores": "DIÁRIO - 3 PESCADORES",
    "Vila Velha": "DIÁRIO - VILA VELHA",
}


def _diario_find_col(header, start, end, keyword, exclude):
    for c in range(start, end):
        if c in exclude:
            continue
        h = _norm_txt(header[c]) if c < len(header) else ""
        if keyword in h:
            return c
    return None


def _diario_parse_block_cols(header, start, end):
    claimed = set()
    col_data = _diario_find_col(header, start, end, "DATA", claimed)
    if col_data is not None:
        claimed.add(col_data)
    col_morador = _diario_find_col(header, start, end, "MORADOR", claimed)
    if col_morador is not None:
        claimed.add(col_morador)
    col_mercosul = _diario_find_col(header, start, end, "MERCOSUL", claimed)
    if col_mercosul is not None:
        claimed.add(col_mercosul)
    col_estrangeiro = _diario_find_col(header, start, end, "ESTRANGEIRO", claimed)
    if col_estrangeiro is not None:
        claimed.add(col_estrangeiro)
    col_brasil = _diario_find_col(header, start, end, "BRASIL", claimed)
    if col_brasil is not None:
        claimed.add(col_brasil)
    col_local = _diario_find_col(header, start, end, "LOCAL", claimed)
    return {
        "data": col_data, "morador": col_morador, "mercosul": col_mercosul,
        "estrangeiro": col_estrangeiro, "brasil": col_brasil, "local": col_local,
    }


def _diario_categoria_labels(cols):
    """Decide os nomes de categoria deste bloco a partir de quais colunas existem -- os
    mesmos nomes já usados na aba Mix de Origem mensal, pra herdar cor/legenda no front."""
    labels = []
    if cols["morador"] is not None and cols["local"] is None:
        labels.append(("Morador/Estado", "morador"))
    elif cols["morador"] is not None and cols["local"] is not None:
        labels.append(("Morador", "morador"))
        labels.append(("Estado", "local"))
    elif cols["local"] is not None:
        labels.append(("Local", "local"))
    if cols["brasil"] is not None:
        labels.append(("Brasileiros", "brasil"))
    if cols["mercosul"] is not None:
        labels.append(("Mercosul", "mercosul"))
    if cols["estrangeiro"] is not None:
        labels.append(("Estrangeiros", "estrangeiro"))
    return labels


def _parse_diario_sheet(rows):
    if len(rows) < 3:
        return None
    header = rows[1]
    # BUG EVITADO: a 1a linha (marcador de ano por bloco, tipo "2026"/"2025") não fica
    # alinhada de forma confiável com o inicio real de cada bloco -- em algumas abas vem
    # deslocada 1-2 colunas pra direita, em outras (PNI) o marcador do bloco de 2026 nem
    # existe. Por isso os blocos são identificados pela própria coluna "Data" da 2a linha
    # (cabeçalho de verdade), que sempre existe e sempre abre cada bloco.
    data_cols = [c for c, v in enumerate(header) if "DATA" in _norm_txt(v)]
    if len(data_cols) < 2:
        return None
    c26_start, c25_start = data_cols[0], data_cols[1]
    c25_end = data_cols[2] if len(data_cols) > 2 else len(header)

    cols26 = _diario_parse_block_cols(header, c26_start, c25_start)
    cols25 = _diario_parse_block_cols(header, c25_start, c25_end)
    if cols26["data"] is None or cols25["data"] is None:
        return None

    labels26 = _diario_categoria_labels(cols26)
    if not labels26:
        return None
    categorias = [lbl for lbl, _key in labels26]
    labels25_map = dict(_diario_categoria_labels(cols25))  # label -> key, pra achar a coluna certa no bloco 2025

    dias = []
    for row in rows[2:]:
        data26_serial = cell(row, cols26["data"])
        if data26_serial is None:
            continue
        data26 = serial_to_date(data26_serial)
        if data26 is None:
            continue
        v26, v25 = {}, {}
        for lbl, key in labels26:
            v26[lbl] = cell(row, cols26[key]) if cols26.get(key) is not None else None
        for lbl in categorias:
            key25 = labels25_map.get(lbl)
            v25[lbl] = cell(row, cols25[key25]) if key25 and cols25.get(key25) is not None else None
        # pula dias sem nenhum valor real em 2026 (linhas-molde do resto do ano ainda vazias)
        if all(v is None for v in v26.values()):
            continue
        dias.append({"data": data26.isoformat(), "v26": v26, "v25": v25})

    return {"categorias": categorias, "dias": dias}


def build_mix_origem_diario(service, spreadsheet_id):
    result = {}
    for park, sheet_name in DIARIO_SHEET_NAMES.items():
        try:
            rows = get_values(service, spreadsheet_id, sheet_name)
            parsed = _parse_diario_sheet(rows)
        except Exception as e:
            print(f"AVISO: falha ao ler aba diária de {park} ({e}) -- fica sem dado diário neste ciclo.", file=sys.stderr)
            parsed = None
        if parsed and park == "BioParque":
            # saiu do Grupo Cataratas a partir de Agosto/2026 -- mantém só os dias
            # anteriores (Jan-Jul), pro Mix de Origem acumulado/histórico continuar
            # contando o período em que o parque de fato fazia parte do grupo.
            parsed = {
                "categorias": parsed["categorias"],
                "dias": [d for d in parsed["dias"] if d["data"] < "2026-08-01"],
            }
        if parsed:
            result[park] = parsed
    return result


# ---------------------------------------------------------------------------
# [2026] Mix OBZ e visitação.xlsx -> aba "AQF E M3F | SMorador"
# Captação PNI (com e sem morador) de M3F e AquaFoz -- pedido do usuário (07/08/2026):
# usar direto essa aba de apoio, que já calcula a captação de forma proporcional aos dias
# do mês vigente (em vez de recalcular no painel Realizado/PNI e aplicar um fator fixo).
#
# Layout da aba (mantido manualmente pelo time, um bloco novo por mês, empilhado):
#   linha N       : nome do mês solto na coluna A (ex.: "Julho", depois "Agosto" abaixo)
#   linha N+1     : "M3F" (coluna A) e "Aquafoz" (coluna E) -- cabeçalho de cada metade
#   linha N+3     : "Captação PNI" (rótulo, mesmas colunas do cabeçalho)
#   linha N+4     : o VALOR da Captação PNI (com morador) do mês
#   linha N+5     : "*SEM MORADOR:" (rótulo) + o valor logo na coluna seguinte
# O bloco ANUAL (M3F/Aquafoz "ANO", sempre o acumulado mais recente) fica ao lado do
# PRIMEIRO bloco mensal, achado pelo rótulo exato "ANO" (blocos antigos ficam com o
# rótulo "Anual até <mês>" e são ignorados -- o time só marca visualmente como obsoleto,
# sem apagar, então não dá pra confiar em posição, só no texto exato).
#
# BUG EVITADO: a versão anterior lia por POSIÇÃO FIXA (linha 5/6, colunas 0/1/4/5) --
# funcionava só enquanto existisse um único bloco "mês corrente" na aba. A partir do
# momento em que o time começou a empilhar um bloco novo por mês (Julho, Agosto, ...)
# pra manter histórico, a leitura por posição fixa ia sempre nos mesmos números antigos
# (Julho), nunca no mês novo -- e não tinha nenhum jeito de saber qual bloco escolher sem
# procurar pelo nome do mês. Agora cada bloco é achado pelo próprio rótulo.
# ---------------------------------------------------------------------------

def _norm_park_semmorador(nome):
    n = str(nome or "").strip().lower()
    if "aqua" in n and "foz" in n:
        return "AquaFoz"
    if n == "m3f":
        return "M3F"
    return None


def build_semmorador_ratio(service, spreadsheet_id, sheet_name):
    rows = get_values(service, spreadsheet_id, sheet_name)
    resultado = {"M3F": {"porMes": {}, "ano": None}, "AquaFoz": {"porMes": {}, "ano": None}}

    # --- Blocos MENSAIS (um por mês, achado pelo nome do mês solto na coluna A) ---
    for i, row in enumerate(rows):
        mes_label = str(cell(row, 0) or "").strip()
        if mes_label not in MESES_PT:
            continue
        header_row = rows[i + 1] if i + 1 < len(rows) else []
        # valida que é mesmo um bloco de Captação (evita casar com "Julho"/"Agosto"
        # soltos em qualquer outro lugar da aba, que tem mais de 1000 linhas)
        if not any(_norm_park_semmorador(v) for v in header_row[:16]):
            continue
        captacao_row = rows[i + 4] if i + 4 < len(rows) else []
        semmorador_row = rows[i + 5] if i + 5 < len(rows) else []
        for c, v in enumerate(header_row[:16]):
            park = _norm_park_semmorador(v)
            # BUG EVITADO: o bloco ANUAL ("M3F"/"Aquafoz" de novo, mais a direita) fica na
            # MESMA linha do cabeçalho do primeiro bloco mensal (a planilha desenha os dois
            # lado a lado) -- sem essa guarda, a segunda ocorrência de "M3F"/"Aquafoz" (a do
            # bloco ANUAL) sobrescreveria o valor mensal certo com o valor anual errado.
            # Como a coluna do bloco mensal sempre vem ANTES (esquerda) da coluna do bloco
            # anual, ficar com a PRIMEIRA ocorrência resolve sem precisar fixar coluna.
            if park is None or mes_label in resultado[park]["porMes"]:
                continue
            captacao = cell(captacao_row, c)
            if not isinstance(captacao, (int, float)):
                continue
            semmorador = cell(semmorador_row, c + 1)
            resultado[park]["porMes"][mes_label] = {
                "captacao": captacao,
                "semMorador": semmorador if isinstance(semmorador, (int, float)) else None,
            }

    # --- Bloco ANUAL atual (rótulo exato "ANO" -- ignora "Anual até <mês>", obsoleto) ---
    for i, row in enumerate(rows[:15]):
        for c, v in enumerate(row[:20]):
            if str(v or "").strip() != "ANO":
                continue
            header_row = rows[i - 6] if i - 6 >= 0 else []
            park = _norm_park_semmorador(cell(header_row, c))
            if park is None or resultado[park]["ano"] is not None:
                continue
            captacao_row = rows[i - 3] if i - 3 >= 0 else []
            semmorador_row = rows[i - 2] if i - 2 >= 0 else []
            captacao = cell(captacao_row, c)
            if not isinstance(captacao, (int, float)):
                continue
            semmorador = cell(semmorador_row, c + 1)
            resultado[park]["ano"] = {
                "captacao": captacao,
                "semMorador": semmorador if isinstance(semmorador, (int, float)) else None,
            }
    return resultado


# ---------------------------------------------------------------------------
# Share E-commerce_Parques Rio - _2026.xlsx -> aba "Dash Share GC"
# ---------------------------------------------------------------------------

def build_dash_share_gc(service, spreadsheet_id, sheet_name):
    rows = get_values(service, spreadsheet_id, sheet_name)
    periodo = cell(rows[0], 1)
    nomes = {"Aquario": 4, "Bioparque": 5, "Paineiras": 6, "Marco": 7, "Aquafoz": 8, "Grupo Cataratas": 10}
    parques = []
    for nome, ridx in nomes.items():
        r = rows[ridx]
        parques.append({
            "parque": nome,
            "metaAnual": cell(r, 1),
            "ecommerce2026": cell(r, 2),
            "visitacao2026": cell(r, 3),
            "share2026": cell(r, 4),
            "ecommerce2025": cell(r, 5),
            "visitacao2025": cell(r, 6),
            "share2025": cell(r, 7),
            "deltaPP": cell(r, 8),
            "gapMeta": cell(r, 9),
        })
    return {"periodo": periodo, "parques": parques}


# ---------------------------------------------------------------------------
# INVESTIMENTO MARKETING _ 2026.xlsx -> aba "acompanhamento mkt"
# ---------------------------------------------------------------------------

BLOCK_NAME_MAP = {
    "AQUARIO": "AquaRio", "BIOPARQUE": "BioParque", "PAINEIRAS": "Paineiras",
    "AQUAFOZ": "AquaFoz", "M3F": "M3F", "VILA VELHA": "Vila Velha",
    "3 PESCADORES": "Três Pescadores",
}


def build_invest_mkt_resumo(service, spreadsheet_id, sheet_name):
    rows = get_values(service, spreadsheet_id, sheet_name)
    resumo = {}
    for i, row in enumerate(rows):
        for c, val in enumerate(row):
            if val in BLOCK_NAME_MAP:
                # BUG EVITADO: nome de parque (ex.: "M3F", "PAINEIRAS") pode aparecer em mais
                # de um lugar na aba (ex.: legenda, celula de referencia). So' tratamos como
                # o INICIO de um bloco de verdade se a linha logo abaixo tiver o cabecalho
                # "MES" na mesma coluna -- senao ficamos vulneraveis a sobrescrever um bloco
                # bom com lixo de uma ocorrencia errada do nome mais abaixo na planilha.
                header_below = cell(rows[i + 1], c) if i + 1 < len(rows) else None
                if not (isinstance(header_below, str) and header_below.strip().upper() in ("MÊS", "MES")):
                    continue
                park = BLOCK_NAME_MAP[val]
                resumo[park] = {}
                r = i + 2
                while r < len(rows):
                    mes = cell(rows[r], c)
                    if not mes:
                        break
                    disp = cell(rows[r], c + 1)
                    real = cell(rows[r], c + 2)
                    saldo = cell(rows[r], c + 3)
                    resumo[park][str(mes).strip().upper()] = {
                        "disponivel": disp if disp not in ("", "#N/A") else None,
                        "realizado": real if real not in ("", "#N/A") else None,
                        "saldo": saldo if saldo not in ("", "#N/A") else None,
                    }
                    r += 1
    return resumo


# ---------------------------------------------------------------------------
# INVESTIMENTO MARKETING _ 2026.xlsx -> uma aba por mes (JANEIRO..JULHO), lista
# de campanhas/linhas de gasto. O cabecalho MUDA de posicao de mes para mes
# (ex.: "SETOR" vira "CUSTO" em Junho/Julho, "RUNRUN IT" some em alguns meses),
# entao procuramos as colunas pelo NOME do cabecalho em vez de por indice fixo.
# ---------------------------------------------------------------------------

# nomes possiveis de cabecalho -> campo de saida (primeiro que bater, na ordem da linha)
# BUG EVITADO: "SETOR" (area: Performance/Marca & Experiencia/Conteudo/InHouse) e "CUSTO"
# (tipo de custo: Custo Fixo/Midia OFF/Midia ON/Eventos) sao DUAS colunas DIFERENTES que
# coexistem em Junho/Julho -- nao aliases uma da outra. Antes eram tratadas como a mesma
# coisa e a coluna CUSTO era sempre descartada (SETOR vinha primeiro na linha).
DETAIL_HEADER_ALIASES = {
    "parque": ["PARQUE"],
    "setor": ["SETOR"],
    "custo": ["CUSTO"],
    "fornecedor": ["FORNECEDOR"],
    "descricao": ["DESCRIÇAO DO SERVIÇO", "DESCRIÇÃO DO SERVIÇO"],
    "runrun": ["RUNRUN IT"],
    "valor": ["VALOR"],
    "mesCompetencia": ["MêS DE COMPETÊNCIA", "MÊS DE COMPETÊNCIA"],
    "observacao": ["OBSERVAÇÃO"],
}


def _find_detail_columns(header_row):
    cols = {}
    for c, raw in enumerate(header_row):
        if not raw:
            continue
        label = str(raw).strip().upper()
        for field, aliases in DETAIL_HEADER_ALIASES.items():
            if field in cols:
                continue
            if label in [a.upper() for a in aliases]:
                cols[field] = c
    return cols


def _parse_valor(v):
    """VALOR vem ora como numero, ora como texto 'R$ 1.780,00' -- normaliza pros dois casos."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("R$", "").strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _clean_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Plano de Mídia - 2026.xlsx (uma aba por parque/mês, nomeada à mão pelo time de mídia --
# SEM padrão fixo, diferente de Investimento Marketing que usa JANEIRO..DEZEMBRO). Por isso
# não dá pra montar o nome da aba dinamicamente a partir do mês -- este mapa precisa ser
# atualizado manualmente conforme novas abas forem criadas (mesmo espírito de
# MARKETING_ACTIONS, mais acima). Quando um parque tem mais de uma aba no mesmo mês (ex.:
# duas verbas/tranches), uso uma lista -- os valores e lançamentos das abas são somados.
# ---------------------------------------------------------------------------
PLANO_MIDIA_SHEETS = {
    "Janeiro":   {"AquaRio": " AQUARIO JAN 2026", "BioParque": "BIO JANEIRO 2026"},
    "Fevereiro": {"AquaRio": "AQUARIO FEV 2026", "BioParque": "BIO fev 2026"},
    "Março":     {"AquaRio": "AQUARIO MAR 2026", "BioParque": "BIO MAR 2026"},
    "Abril":     {"AquaRio": "AQUARIO ABRIL", "BioParque": "BIOPARQUE ABRIL", "Vila Velha": "VILA VELHA"},
    "Maio":      {"AquaRio": ["AQUARIO MAIO 1", "AQUARIO MAIO 2"], "BioParque": "BIOPARQUE MAIO", "AquaFoz": "AQUAFOZ MAIO"},
    "Junho":     {"AquaRio": "AQUARIO JUNHO", "BioParque": "BIOPARQUE JUNHO"},
    "Julho":     {"AquaRio": "AQUARIO JULHO", "BioParque": "BIOPARQUE JULHO", "AquaFoz": "AQUAFOZ JULHO",
                  "M3F": "M3F JULHO", "Vila Velha": "VILA VELHA JULHO", "3P": "3 PESCADORES JULHO"},
    "Agosto":    {"AquaRio": "AQUARIO AGOSTO", "BioParque": "BIOPARQUE AGOSTO", "AquaFoz": "AQUAFOZ AGOSTO",
                  "M3F": "M3F AGOSTO", "Vila Velha": "VILA VELHA AGOSTO", "3P": "3 PESCADORES AGOSTO"},
}
PLANO_MIDIA_PARKS = ["AquaRio", "BioParque", "AquaFoz", "M3F", "Vila Velha", "3P"]


def _parse_plano_midia_sheet(rows):
    """Recebe as linhas (get_values) de UMA aba de parque/mês do Plano de Mídia e devolve
    {"budget", "previsto", "saldo", "items"}. O layout de colunas varia de parque pra parque
    (ex.: 'Empresa' vs 'Fornecedor' vs 'Veículo', com/sem coluna 'Card') e a linha de
    cabeçalho também varia de posição (2 ou 3, conforme tem ou não uma linha de observação no
    meio) -- por isso acha tudo pelo rótulo em vez de índice fixo, procurando a célula
    "Parque" na coluna A pra achar o cabeçalho.

    Nem toda aba tem linha-resumo ("ORÇADO"/"PREVISTO"/"SALDO"/"TOTAL") -- e nas abas mais
    antigas (Jan a Mar) a lista de lançamentos é uma lista de PROPOSTAS/opções em negociação,
    não só o que foi fechado, então o valor total dos lançamentos pode somar bem mais que o
    orçamento (várias alternativas concorrentes pro mesmo espaço). Por isso "previsto" NUNCA
    é calculado somando os lançamentos -- só vem de uma linha-resumo explícita da própria
    planilha (ORÇADO/PREVISTO/TOTAL). Sem uma dessas linhas, "previsto" e "saldo" ficam None
    (sem dado) em vez de arriscar um número que pode estar bem errado.
    """
    budget = cell(rows[0], 1) if rows else None
    budget = budget if isinstance(budget, (int, float)) else None

    header_idx = None
    for i in range(min(8, len(rows))):
        if _norm_txt(cell(rows[i], 0)) == "PARQUE":
            header_idx = i
            break
    if header_idx is None:
        return {"budget": budget, "previsto": None, "saldo": None, "items": []}

    header = rows[header_idx]
    col_map = {c: _norm_txt(v) for c, v in enumerate(header) if _norm_txt(v)}

    def find_col(*keywords):
        for c, h in col_map.items():
            for kw in keywords:
                if kw in h:
                    return c
        return None

    col_valor = find_col("VALOR")
    col_tipo = find_col("TIPO")
    col_desc = find_col("PROPRIEDADE")
    col_forn = find_col("EMPRESA", "FORNECEDOR", "VEICULO")
    col_status = find_col("STATUS")
    col_ok = find_col("OK")

    items = []
    orcado_val = None
    previsto_val = None
    saldo_val = None
    blank_streak = 0
    i = header_idx + 1
    maxi = min(len(rows), header_idx + 60)
    while i < maxi:
        row = rows[i]
        row_norm = [_norm_txt(v) for v in row]
        nums = [v for v in row if isinstance(v, (int, float))]
        if "ORCADO" in row_norm or "TOTAL" in row_norm:
            if nums: orcado_val = nums[0]
            blank_streak = 0; i += 1; continue
        if "PREVISTO" in row_norm:
            if nums: previsto_val = nums[0]
            blank_streak = 0; i += 1; continue
        if "SALDO" in row_norm:
            if nums: saldo_val = nums[0]
            blank_streak = 0; i += 1; continue
        # Só conta como lançamento real se a coluna "Parque" (a primeira) repetir o nome do
        # parque, igual em toda linha de item de verdade nessas planilhas -- isso evita
        # "vazar" pra dentro da lista notas soltas tipo "CONTATO FORNECEDORES AQUI" ou uma
        # tabela de recorte por parque que aparece depois da lista em algumas abas (ex.:
        # " AQUARIO JAN 2026"), que não têm nada na coluna A mas por coincidência de posição
        # de coluna acabam caindo em cima de "Tipo"/"Status"/"OK?".
        # (algumas abas -- ex. a genérica "VILA VELHA", que empilha Abril e Maio na mesma
        # aba -- repetem um segundo cabeçalho "Parque/Tipo/.../STATUS" mais pra baixo; sem
        # esse segundo `or`, esse cabeçalho repetido também "vazaria" pra lista de itens).
        if not _clean_str(cell(row, 0)) or _norm_txt(cell(row, 0)) == "PARQUE":
            blank_streak += 1
            if blank_streak >= 25:
                break
            i += 1
            continue
        valor = cell(row, col_valor) if col_valor is not None else None
        valor = valor if isinstance(valor, (int, float)) else None
        desc = _clean_str(cell(row, col_desc)) if col_desc is not None else None
        forn = _clean_str(cell(row, col_forn)) if col_forn is not None else None
        tipo = _clean_str(cell(row, col_tipo)) if col_tipo is not None else None
        ok = _clean_str(cell(row, col_ok)) if col_ok is not None else None
        if valor is not None or desc or forn or tipo:
            items.append({
                "tipo": tipo,
                "fornecedor": forn,
                "descricao": desc,
                "valor": valor,
                "status": _clean_str(cell(row, col_status)) if col_status is not None else None,
                "ok": ok,
            })
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= 25:
                break
        i += 1

    # "Previsto" (comprometido) só conta os lançamentos com "OK?" = "Sim" -- qualquer outro
    # status ("Não", "Análise Camila", em branco etc.) é só proposta/opção em negociação e não
    # deve entrar na soma que aparece nos cards. Isso substitui a leitura das linhas-resumo
    # ORÇADO/PREVISTO/TOTAL da própria planilha (que não davam pra confiar que seguiam essa
    # mesma regra) -- só cai de volta pra elas se a aba nem tiver uma coluna "OK?".
    if col_ok is not None:
        previsto = sum(it["valor"] for it in items if it["valor"] is not None and _norm_txt(it["ok"]) == "SIM")
    else:
        previsto = previsto_val if previsto_val is not None else orcado_val

    if budget is None and previsto is not None and saldo_val is not None:
        budget = previsto + saldo_val
    saldo = (budget - previsto) if (budget is not None and previsto is not None) else saldo_val

    return {"budget": budget, "previsto": previsto, "saldo": saldo, "items": items}


def build_plano_midia(service, spreadsheet_id):
    """Lê a planilha 'Plano de Mídia' -- cada parque/mês tem sua própria aba oculta (ver
    PLANO_MIDIA_SHEETS acima). Diferente de Investimento Marketing, aqui NÃO há comparativo
    com 2025: é só o planejado do ano corrente (Orçado = verba do mês, Previsto = já
    comprometido, Saldo = o que resta). Cada aba é lida isoladamente (try/except) -- se uma
    não existir mais ou vier com erro, fica de fora sem derrubar o resto do mês/pipeline.
    """
    resumo = {}
    detail = {}
    for mes, parques_sheets in PLANO_MIDIA_SHEETS.items():
        resumo[mes] = {}
        detail[mes] = []
        for park in PLANO_MIDIA_PARKS:
            if park == "BioParque" and mes in BIOPARQUE_MESES_PT_POS_SAIDA:
                continue  # saiu do Grupo Cataratas -- some do Plano de Mídia a partir daqui
            sheet_names = parques_sheets.get(park)
            if not sheet_names:
                continue
            if isinstance(sheet_names, str):
                sheet_names = [sheet_names]
            budget_total = 0.0
            previsto_total = 0.0
            saldo_total = 0.0
            has_budget = False
            has_previsto = False
            has_saldo = False
            for sheet_name in sheet_names:
                try:
                    rows = get_values(service, spreadsheet_id, sheet_name)
                    parsed = _parse_plano_midia_sheet(rows)
                except Exception as e:
                    print(f"AVISO: falha ao ler Plano de Mídia '{sheet_name}' ({e}) -- aba ignorada.", file=sys.stderr)
                    continue
                if parsed["budget"] is not None:
                    budget_total += parsed["budget"]; has_budget = True
                if parsed["previsto"] is not None:
                    previsto_total += parsed["previsto"]; has_previsto = True
                if parsed["saldo"] is not None:
                    saldo_total += parsed["saldo"]; has_saldo = True
                for it in parsed["items"]:
                    detail[mes].append({**it, "parque": park})
            resumo[mes][park] = {
                "orcado": budget_total if has_budget else None,
                "previsto": previsto_total if has_previsto else None,
                "saldo": saldo_total if has_saldo else None,
            }
    return {"resumo": resumo, "detail": detail}


def build_invest_mkt_detail(service, spreadsheet_id, meses):
    """Le a lista de campanhas de cada aba mensal. Retorna {mes: [ {parque, setor,
    custo, fornecedor, descricao, valor, observacao}, ... ]}.

    Lê TODOS os meses passados (não só os "meses_com_dados") -- a aba mensal de
    Investimento Marketing já recebe lançamentos de custo fixo com antecedência (ex.:
    Agosto já tinha linhas preenchidas em Julho), então mesmo mês "futuro" pode ter
    detalhe real pra mostrar, mesmo que o resumo geral daquele mês ainda esteja incompleto.
    Cada mês é lido isoladamente (try/except) -- se uma aba não existir ou vier vazia,
    esse mês fica com lista vazia, sem derrubar os outros meses nem o resto do pipeline.
    """
    detail = {}
    for mes in meses:
        try:
            rows = get_values(service, spreadsheet_id, mes)
        except Exception as e:
            print(f"AVISO: falha ao ler detalhe de Investimento Marketing de {mes} ({e}) -- mês fica vazio.", file=sys.stderr)
            detail[mes] = []
            continue
        if not rows:
            detail[mes] = []
            continue
        cols = _find_detail_columns(rows[0])
        if "parque" not in cols or "valor" not in cols:
            detail[mes] = []
            continue
        items = []
        for row in rows[1:]:
            parque = cell(row, cols.get("parque"))
            if not parque:
                continue
            items.append({
                "parque": str(parque).strip(),
                "setor": _clean_str(cell(row, cols.get("setor"))) if "setor" in cols else None,
                "custo": _clean_str(cell(row, cols.get("custo"))) if "custo" in cols else None,
                "fornecedor": cell(row, cols.get("fornecedor")) if "fornecedor" in cols else None,
                "descricao": cell(row, cols.get("descricao")) if "descricao" in cols else None,
                "valor": _parse_valor(cell(row, cols.get("valor"))),
                "observacao": cell(row, cols.get("observacao")) if "observacao" in cols else None,
            })
        detail[mes] = items
    return detail


MESES_NOME_PT = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "marco": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12,
}


# ---------------------------------------------------------------------------
# Visitação Parques 2026.xlsx -> aba "APOIO REPORT"
# Aba de apoio usada hoje (fora deste painel) para alimentar o gerador de Report Diário
# (Gemini/Claude configurado com os "Gold Standards" da equipe). O objetivo aqui NAO e'
# gerar o texto do report (isso continua no agente ja calibrado) -- e' so' automatizar a
# parte chata/manual: ler direto da planilha os dados que hoje sao copiados a mao.
# Layout (bloco "Realizado ontem", colunas B a E): marcador de regiao solto na coluna A
# em alguma linha dentro do bloco de cada regiao (ex.: "Parques Rio", "Parques Foz",
# "Soul Parques"); nome do parque na coluna C seguido de ate' 5 linhas de metrica
# ("Realizado 2026"/"OBZ 2026"/"% OBZ"/"Realizado 2025"/"% 2025") na coluna C com valor na
# coluna D. AquaFoz so' tem 3 linhas de metrica (sem Realizado/% 2025 -- consistente com a
# regra "AquaFoz nunca tem comparativo vs 2025", ja' embutida na propria planilha).
# Metricas extras (captacao/share dos parques satelite) aparecem soltas, ou na coluna C
# com valor na mesma linha coluna D (ex.: "Captação CV"), ou na coluna E com o valor na
# MESMA coluna uma linha abaixo (ex.: "Share PNI (parcial)") -- copiado fielmente, sem
# tentar renomear/reinterpretar o rotulo.
# Bloco "Agosto(PARCIAL)" (resumo mensal, colunas G-L): tabela reta por parque + linhas de
# agregado (Grupo Cataratas SSS / Total Grupo Cataratas / Total Soul Parques / Total Geral).
# NAO extraidos nesta versao (layout ainda ambiguo ou vazio nesta planilha-exemplo, a
# confirmar com o usuario antes de automatizar): bloco "Clima de Ontem" (sem valores
# preenchidos na planilha lida), bloco "Realizado Final de Semana" (erros #DIV/0! nesta
# leitura -- so' faz sentido numa segunda-feira real) e bloco "META OBZ (Julho)" (colunas
# de valor sem cabecalho claro).
# ---------------------------------------------------------------------------

APOIO_REPORT_METRICAS = ["Realizado 2026", "OBZ 2026", "% OBZ", "Realizado 2025", "% 2025"]

# BUG EVITADO: o marcador de regiao da coluna A ("Parques Rio"/"Parques Foz"/"Soul Parques")
# vem de uma celula mesclada no Google Sheets -- ao ler via openpyxl/API, o valor so'
# aparece na linha exata onde a mesclagem "ancora" (uma linha arbitraria dentro do bloco,
# nao necessariamente a primeira), nunca em todas as linhas que ela cobre visualmente.
# Tentar usar essa coluna pra decidir a regiao de cada parque quebra (ex.: BioParque cai
# sem regiao porque o marcador so' aparece 2 linhas depois, na propria linha de "OBZ
# 2026"). Por isso a regiao de cada parque vem de uma lista fixa (mesmo padrao ja' usado
# em MIX_ORIGEM_PARK_NAMES/DIARIO_SHEET_NAMES pra outras abas com nomes/posicoes
# inconsistentes), nao da leitura da coluna A.
APOIO_REPORT_PARK_REGIAO = {
    "AquaRio": "Parques Rio", "BioParque": "Parques Rio", "Paineiras": "Parques Rio",
    "PNI": "Parques Foz", "URBIA + CATARATAS (PNI)": "Parques Foz",
    "M3F": "Parques Foz", "AquaFoz": "Parques Foz",
    "Três Pescadores": "Soul Parques", "Vila Velha": "Soul Parques",
}
APOIO_REPORT_PARK_ALIAS = {"URBIA + CATARATAS (PNI)": "PNI"}


def _apoio_report_realizado_ontem(rows, linha_inicio, linha_fim):
    """Percorre o bloco 'Realizado ontem' (colunas C=2 a E=4) e devolve
    {regiao: {parque: {"metricas": {...}, "extras": {...}}}}. Ve' Nota acima sobre por que
    a regiao vem de APOIO_REPORT_PARK_REGIAO, nao da coluna A."""
    resultado = {}
    parque_atual = None
    regiao_atual = None
    for r in range(linha_inicio, linha_fim):
        row = rows[r]
        label_c = cell(row, 2)
        valor_d = cell(row, 3)
        if isinstance(label_c, str) and label_c.strip():
            label_c = label_c.strip()
            if label_c in APOIO_REPORT_METRICAS:
                if parque_atual is not None:
                    resultado[regiao_atual][parque_atual]["metricas"][label_c] = valor_d
            elif label_c in APOIO_REPORT_PARK_REGIAO:
                parque_atual = APOIO_REPORT_PARK_ALIAS.get(label_c, label_c)
                regiao_atual = APOIO_REPORT_PARK_REGIAO[label_c]
                resultado.setdefault(regiao_atual, {})
                resultado[regiao_atual].setdefault(parque_atual, {"metricas": {}, "extras": {}})
            elif parque_atual is not None:
                # rotulo "extra" (captacao/share) com valor na propria linha coluna D
                resultado[regiao_atual][parque_atual]["extras"][label_c] = valor_d

        # rotulos "extra" que vivem na coluna E (ex.: "Share PNI (parcial)"), com o valor
        # uma linha abaixo, na mesma coluna E -- padrao observado na planilha real.
        label_e = cell(row, 4)
        if isinstance(label_e, str) and label_e.strip() and parque_atual is not None:
            valor_e = cell(rows[r + 1], 4) if r + 1 < linha_fim else None
            resultado[regiao_atual][parque_atual]["extras"][label_e.strip()] = valor_e
    return resultado


def _apoio_report_resumo_mensal(rows, nome_col, valor_col_inicio, linha_titulo, linha_fim):
    """Le' o bloco reto 'Agosto(PARCIAL)' (ou equivalente): titulo do mes uma coluna a
    direita do nome do parque, na linha do cabecalho; depois uma linha por parque/agregado
    com Realizado/OBZ/%OBZ/2025/%2025."""
    titulo = cell(rows[linha_titulo], valor_col_inicio)
    header_row = rows[linha_titulo + 1]
    headers = [cell(header_row, valor_col_inicio + i) for i in range(5)]
    parques = {}
    for r in range(linha_titulo + 2, linha_fim):
        nome = cell(rows[r], nome_col)
        if not isinstance(nome, str) or not nome.strip():
            continue
        valores = {}
        for i, h in enumerate(headers):
            if isinstance(h, str):
                chave = _clean_str(h)
            elif isinstance(h, float) and h.is_integer():
                chave = str(int(h))
            else:
                chave = str(h)
            valores[chave or f"col{i}"] = cell(rows[r], valor_col_inicio + i)
        parques[nome.strip()] = valores
    return {"titulo": _clean_str(titulo) if isinstance(titulo, str) else titulo, "parques": parques}


def build_apoio_report(service, spreadsheet_id, sheet_name):
    """Aba 'APOIO REPORT': fonte primaria hoje usada (fora deste painel) pelo gerador de
    Report Diario. So' automatiza a LEITURA -- a geracao do texto continua no
    Gemini/Claude ja calibrado com os Gold Standards da equipe (ver documentos anexados)."""
    rows = get_values(service, spreadsheet_id, sheet_name)
    n = len(rows)

    realizado_ontem = _apoio_report_realizado_ontem(rows, 0, min(66, n))

    resumo_mensal = {}
    for r in range(0, min(20, n)):
        titulo = cell(rows[r], 7)
        if isinstance(titulo, str) and titulo.strip() and cell(rows[r + 1], 6) == "GRUPO CATARATAS":
            resumo_mensal = _apoio_report_resumo_mensal(rows, 6, 7, r, r + 14)
            break

    # BioParque saiu do Grupo Cataratas a partir de Agosto/2026 -- o Report é sempre sobre
    # "ontem"/"o mês em curso", que hoje já é sempre Agosto/2026 em diante, então o parque
    # não deve mais aparecer aqui (mesmo que a planilha ainda tenha a linha preenchida).
    # Continua reconhecido em APOIO_REPORT_PARK_REGIAO só pra consumir corretamente o bloco
    # dele na leitura (sem "vazar" pro parque vizinho) -- é descartado só no final, aqui.
    realizado_ontem.get("Parques Rio", {}).pop("BioParque", None)
    resumo_mensal.get("parques", {}).pop("BioParque", None)

    return {
        "realizadoOntem": realizado_ontem,
        "resumoMensal": resumo_mensal,
    }


def _eventos_ano_com_rollover(mes_fim, data_ini):
    """Se o mes final extraido do texto for menor que o mes da data-ancora, assume que o
    intervalo atravessa a virada de ano (ex.: 28/12 a 05/01) -- rollover pro ano seguinte."""
    return data_ini.year + (1 if mes_fim < data_ini.month else 0)


def _extrair_data_fim(texto, data_ini):
    """Extrai (best-effort) a data de termino de um evento a partir do texto livre da coluna
    'Evento' -- a planilha nao tem coluna de Data Fim, o prazo sempre vem embutido na frase
    (ex.: "Feriado (18/04 até 24/04)", "de 19 a 22 de junho", "Período 09/07 a 12/06").
    Tenta alguns padroes comuns, na ordem: (1) duas datas completas DD/MM/AAAA; (2) "a partir
    do dia DD/MM ... até DD/MM"; (3) "DD/MM (até|a) DD/MM"; (4) "de D a D de <mes por extenso>".
    Nao tenta ser exaustivo -- texto livre em portugues nao e' 100% padronizavel, e um humano
    sempre revisa o Report antes de copiar/enviar, entao "pega a maioria dos casos, nunca
    inventa um errado" e' a barra certa aqui, nao "perfeito".
    BUG EVITADO: sem validacao de sanidade, um erro de digitacao na planilha (ex.: "Período
    09/07 a 12/06" quando a data-ancora e' 09/07 -- devia ser 12/07) faria o rollover de ano
    "corrigir" errado pra quase 1 ano depois (12/06 do ano seguinte). Por isso descarta
    qualquer resultado anterior a data_ini ou com mais de 60 dias de distancia -- nesses casos
    cai no fallback seguro (evento de 1 dia so, ancorado em "data").
    """
    if not texto:
        return None
    t = str(texto)
    fim = None

    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})\s*(?:e|a|at[eé])\s*(\d{1,2})/(\d{1,2})/(\d{4})', t, re.I)
    if m:
        dd, mm, yyyy = int(m.group(4)), int(m.group(5)), int(m.group(6))
        try:
            fim = datetime.date(yyyy, mm, dd)
        except ValueError:
            fim = None

    if fim is None:
        m = re.search(r'a\s*partir\s*d[oa]?\s*dia\s*(\d{1,2})/(\d{1,2}).{0,40}?at[eé]\s*(\d{1,2})/(\d{1,2})', t, re.I)
        if m:
            dd2, mm2 = int(m.group(3)), int(m.group(4))
            ano = _eventos_ano_com_rollover(mm2, data_ini)
            try:
                fim = datetime.date(ano, mm2, dd2)
            except ValueError:
                fim = None

    if fim is None:
        m = re.search(r'(\d{1,2})/(\d{1,2})\s*(?:at[eé]|a)\s*(\d{1,2})/(\d{1,2})(?!/\d)', t, re.I)
        if m:
            dd2, mm2 = int(m.group(3)), int(m.group(4))
            ano = _eventos_ano_com_rollover(mm2, data_ini)
            try:
                fim = datetime.date(ano, mm2, dd2)
            except ValueError:
                fim = None

    if fim is None:
        m = re.search(r'\bde\s+(\d{1,2})\s+a\s+(\d{1,2})\s+de\s+(\w+)', t, re.I)
        if m:
            mes = MESES_NOME_PT.get(m.group(3).lower())
            if mes:
                dd2 = int(m.group(2))
                ano = _eventos_ano_com_rollover(mes, data_ini)
                try:
                    fim = datetime.date(ano, mes, dd2)
                except ValueError:
                    fim = None

    if fim is not None and (fim < data_ini or (fim - data_ini).days > 60):
        return None
    return fim


def build_eventos(service, spreadsheet_id, sheet_name):
    """Aba 'Eventos' de Visitação Parques 2026: calendário simples (uma linha por evento) com
    4 colunas -- Data | Parque | Categoria | Evento -- cobrindo Abril/2025 em diante (sem
    separação por aba/mês como as outras abas dessa planilha). A coluna Categoria (adicionada
    depois da 1a versão desta função) é uma das 5 fixas: "Operação do Parque", "Ações
    Comerciais e Institucionais", "Clima e Meio Ambiente", "Calendário Turístico", "Eventos
    Externos". Retorna lista ordenada por data, com ano/mês (PT) já calculados pra facilitar o
    filtro no front-end, mais "dataFim" (best-effort, ver _extrair_data_fim) pra uso do
    gerador de Report -- não existe coluna de Data Fim na planilha, então quando o texto não
    tem um padrão reconhecível "dataFim" fica igual a "data" (evento de 1 dia). O campo
    "parque" é mantido como veio da planilha (ex.: "Aquario", "Parques Rio", "Todos os
    Parques") -- os nomes lá não seguem exatamente a lista canônica de parques do painel,
    então não tentamos remapear.
    """
    rows = get_values(service, spreadsheet_id, sheet_name)
    eventos = []
    for row in rows[1:]:  # linha 0 = cabeçalho (Data / Parque / Categoria / Evento)
        serial = cell(row, 0)
        evento = _clean_str(cell(row, 3))
        if serial is None or not evento:
            continue
        d = serial_to_date(serial)
        if d is None:
            continue
        data_fim = _extrair_data_fim(evento, d) or d
        eventos.append({
            "data": d.isoformat(),
            "dataFim": data_fim.isoformat(),
            "ano": d.year,
            "mes": MESES_PT[d.month - 1],
            "parque": _clean_str(cell(row, 1)) or "",
            "categoria": _clean_str(cell(row, 2)) or "",
            "evento": evento,
        })
    eventos.sort(key=lambda e: e["data"])
    return eventos


# ---------------------------------------------------------------------------
# Visitação Parques HISTORICO_Grupo Cataratas.xlsx (uma aba por parque: AquaRio,
# BioParque, Paineiras, PNI, M3F, AquaFoz). Cada aba tem até 3 blocos empilhados na coluna
# B: o nome do parque (Visitação, ano a ano), "Share " (share %, só em AquaRio/BioParque/
# Paineiras) e "Visão semestral" (2 colunas: Jan-Jun / Jul-Dez). "Consolidado" (soma dos 6
# parques) é calculado aqui, não existe pronto na planilha.
# ---------------------------------------------------------------------------

HISTORICO_PARKS = ["AquaRio", "BioParque", "Paineiras", "PNI", "M3F", "AquaFoz"]
HISTORICO_PARKS_COM_SHARE = ["AquaRio", "BioParque", "Paineiras"]


def _historico_find_row(rows, label):
    """Acha a linha cujo rótulo (coluna B) bate com `label` (ex.: nome do parque, "Share ",
    "Visão semestral") -- ignora espaços sobrando, que a planilha tem em alguns rótulos."""
    target = label.strip().lower()
    for i, row in enumerate(rows):
        v = cell(row, 1)
        if isinstance(v, str) and v.strip().lower() == target:
            return i
    return None


def _historico_parse_block(rows, start_row, ncols):
    """A partir da linha de rótulo do bloco (`start_row`), lê as linhas seguintes no formato
    Ano | valor x ncols (coluna B = ano, colunas seguintes = Jan..Dez ou Sem1/Sem2), até achar
    uma linha sem ano válido na coluna B. Tolera UMA linha de sub-cabeçalho logo no início
    (ex.: "Jan-Jun"/"Jul-Dez" antes dos anos, no bloco Visão Semestral de Paineiras)."""
    anos = {}
    if start_row is None:
        return anos
    r = start_row + 1
    permitiu_subcabecalho = False
    while r < len(rows):
        row = rows[r]
        ano_raw = cell(row, 1)
        if ano_raw is None or ano_raw == "":
            if not anos and not permitiu_subcabecalho and any(
                isinstance(cell(row, 2 + i), str) for i in range(ncols)
            ):
                permitiu_subcabecalho = True
                r += 1
                continue
            break
        try:
            ano = int(float(ano_raw))
        except (TypeError, ValueError):
            break
        valores = [
            cell(row, 2 + i) if isinstance(cell(row, 2 + i), (int, float)) else None
            for i in range(ncols)
        ]
        anos[str(ano)] = valores
        r += 1
    return anos


def build_historico(service, spreadsheet_id):
    parques = {}
    for park in HISTORICO_PARKS:
        rows = get_values(service, spreadsheet_id, park)
        idx_park = _historico_find_row(rows, park)
        anos_visitacao = _historico_parse_block(rows, idx_park, 12)

        idx_share = _historico_find_row(rows, "Share") if park in HISTORICO_PARKS_COM_SHARE else None
        anos_share = _historico_parse_block(rows, idx_share, 12) if idx_share is not None else {}

        idx_sem = _historico_find_row(rows, "Visão semestral")
        anos_semestral = _historico_parse_block(rows, idx_sem, 2) if idx_sem is not None else {}

        if park == "BioParque" and "2026" in anos_visitacao:
            # saiu do Grupo Cataratas a partir de Agosto/2026 -- zera Ago-Dez/2026 (índices
            # 7-11, Jan=0) tanto na visitação quanto no share; Jan-Jul/2026 e todo o
            # histórico anterior (2023-2025) continuam intactos, sem nenhuma alteração.
            for i in range(7, 12):
                if i < len(anos_visitacao["2026"]):
                    anos_visitacao["2026"][i] = None
            if "2026" in anos_share:
                for i in range(7, 12):
                    if i < len(anos_share["2026"]):
                        anos_share["2026"][i] = None

        parques[park] = {
            "anos": anos_visitacao,
            "anosShare": anos_share,
            "anosSemestral": anos_semestral,
        }

    # "Consolidado": soma mês a mês dos 6 parques por ano -- trata parque sem dado naquele
    # mês como 0 na soma (mesma convenção da aba "Grupo Cataratas" da planilha: um parque
    # que ainda não existia naquele ano simplesmente não entra na soma), mas o mês fica
    # None se NENHUM dos 6 parques tiver dado.
    def _consolida(campo, ncols):
        todos_anos = set()
        for p in HISTORICO_PARKS:
            todos_anos.update(parques[p][campo].keys())
        out = {}
        for ano in todos_anos:
            soma = []
            for m in range(ncols):
                valores_mes = [parques[p][campo].get(ano, [None] * ncols)[m] for p in HISTORICO_PARKS]
                validos = [v for v in valores_mes if v is not None]
                soma.append(sum(validos) if validos else None)
            out[ano] = soma
        return out

    parques["Consolidado"] = {
        "anos": _consolida("anos", 12),
        "anosShare": {},  # share é sempre por parque -- não faz sentido "somar" percentuais
        "anosSemestral": _consolida("anosSemestral", 2),
    }

    return {"parques": parques, "parquesComShare": HISTORICO_PARKS_COM_SHARE}


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    with open(args.config, encoding="utf-8") as f:
        cfg = json.load(f)

    service = get_drive_service()

    print("Detectando automaticamente até qual mês já tem Realizado preenchido...", file=sys.stderr)
    try:
        cfg["meses_com_dados"] = detect_meses_com_dados(service, cfg["visitacao_parques_id"], cfg["meses"])
    except Exception as e:
        # Fallback defensivo: se a autodetecção falhar por algum motivo (ex.: planilha
        # temporariamente inacessível), usa o que já estiver salvo no config.json em vez
        # de derrubar o pipeline inteiro.
        print(f"AVISO: falha ao autodetectar meses_com_dados ({e}) -- usando o valor salvo no config.json.", file=sys.stderr)
        cfg.setdefault("meses_com_dados", [])
    print(f"Meses com Realizado detectados: {cfg['meses_com_dados']}", file=sys.stderr)

    print("Lendo Visitação Parques 2026...", file=sys.stderr)
    visitacao = build_visitacao(service, cfg["visitacao_parques_id"], cfg["meses_com_dados"])

    # Meses sem resultado real ainda (Agosto em diante, tipicamente): entram no VISITACAO
    # só com a Meta (projecao "Nova2026") no lugar do OBZ -- Realizado fica None/vazio.
    # TEMPORARIO: usa o dado estatico (VISITACAO_META_ESTATICO) em vez de ler ao vivo do
    # Drive, ate a planilha "Visitação Diária Ponderada" ser compartilhada com a service
    # account. Pra trocar pra leitura ao vivo depois, troque a linha abaixo por:
    #   visitacao_meta = build_visitacao_meta(service, cfg["visitacao_diaria_ponderada_id"], cfg["sheet_names"]["visitacao_diaria_gc_v2"], meses_meta)
    meses_meta = [m for m in cfg["meses"] if m not in cfg["meses_com_dados"]]
    if meses_meta:
        print("Usando Meta estática (Ago-Dez) — planilha ainda não conectada ao Drive...", file=sys.stderr)
        visitacao_meta = VISITACAO_META_ESTATICO
        for mes in meses_meta:
            month_number = MONTH_NUMBER[mes]
            n_days = calendar.monthrange(2026, month_number)[1]
            daily = {}
            for park in cfg["parques"]:
                if park == "BioParque" and not _bioparque_ainda_conta(2026, month_number):
                    continue  # saiu do Grupo Cataratas -- sem Meta projetada daqui em diante
                meta_arr = visitacao_meta.get(mes, {}).get(park)
                daily[park] = {
                    "Realizado 2026": [None] * n_days,
                    "Realizado 2025": [None] * n_days,
                    "OBZ 2026": meta_arr if meta_arr is not None else [None] * n_days,
                }
            visitacao[mes] = {
                "monthNumber": month_number,
                "nDays": n_days,
                "summary": {},
                "daily": daily,
                "atrativos": {"daily": {}, "accum": {}},
            }

    print("Lendo Mapa Clima...", file=sys.stderr)
    clima_emoji = build_clima_emoji(
        service, cfg["visitacao_parques_id"], cfg["sheet_names"]["mapa_clima"], cfg["meses_com_dados"]
    )
    for mes, parques in clima_emoji.items():
        for park, emoji_arr in parques.items():
            if park in visitacao[mes]["daily"]:
                visitacao[mes]["daily"][park]["Emoji"] = emoji_arr

    print("Lendo Captação CV - 3P...", file=sys.stderr)
    cv3p_by_month, cv3p_anual, cv3p_diario = build_captacao_cv_3p(
        service, cfg["visitacao_parques_id"], cfg["sheet_names"]["captacao_cv_3p"]
    )

    print("Lendo Dash Share GC...", file=sys.stderr)
    dash_share_gc = build_dash_share_gc(
        service, cfg["share_ecommerce_id"], cfg["sheet_names"]["dash_share_gc"]
    )

    print("Lendo comparativo proporcional (D-1) do mês vigente...", file=sys.stderr)
    ppt_parcial = {}
    if cfg["meses_com_dados"]:
        try:
            ppt_parcial = build_evolucao_ppt_parcial(
                service, cfg["share_ecommerce_id"], cfg["meses_com_dados"][-1]
            )
        except Exception as e:
            # Não quebra o pipeline se a aba "EVOLUÇÃO<MÊS> (PPT)" ainda não existir ou tiver
            # sido criada com outro nome/layout -- só mantém o comparativo com o mês inteiro
            # de 2025 (comportamento anterior) pra esse mês.
            print(
                f"AVISO: falha ao ler comparativo D-1 do mês vigente ({e}) -- "
                "mantém comparativo com o mês inteiro de 2025.",
                file=sys.stderr,
            )
            ppt_parcial = {}

    print("Lendo Share_Ecommerce_2026 (investimentoMidia)...", file=sys.stderr)
    investimento_midia = build_investimento_midia(
        service, cfg["share_ecommerce_id"], cfg["sheet_names"]["share_ecommerce_2026"],
        cfg["meses_com_dados"], ppt_parcial
    )

    print("Lendo acompanhamento mkt...", file=sys.stderr)
    invest_mkt_resumo = build_invest_mkt_resumo(
        service, cfg["investimento_marketing_id"], cfg["sheet_names"]["acompanhamento_mkt"]
    )

    print("Lendo detalhe de campanhas (Investimento Marketing, mes a mes)...", file=sys.stderr)
    # Le TODOS os meses (nao so' meses_com_dados) -- meses futuros ja podem ter lancamentos
    # de custo fixo cadastrados com antecedencia na planilha (ver comentario na funcao).
    invest_mkt_detail = build_invest_mkt_detail(
        service, cfg["investimento_marketing_id"], cfg["meses"]
    )

    print("Lendo evolução mensal (série histórica)...", file=sys.stderr)
    evolucao_mensal = build_evolucao_mensal(
        service, cfg["share_ecommerce_id"], cfg["sheet_names"]["share_ecommerce_2026"],
        cfg["meses_com_dados"], ppt_parcial
    )

    print("Calculando Share Meta Grupo Cataratas...", file=sys.stderr)
    share_meta_grupo_cataratas = build_share_meta_grupo_cataratas(
        investimento_midia, cfg["meses_com_dados"]
    )

    print("Lendo Captação PNI Sem Morador...", file=sys.stderr)
    semmorador_ratio = build_semmorador_ratio(
        service, cfg["mix_obz_visitacao_id"], cfg["sheet_names"]["smorador"]
    )

    print("Lendo Mix de Origem...", file=sys.stderr)
    mix_origem = build_mix_origem(
        service, cfg["mix_obz_visitacao_id"], cfg["sheet_names"]["mix_origem"]
    )
    mix_origem_acumulado = build_mix_origem(
        service, cfg["mix_obz_visitacao_id"], cfg["sheet_names"]["mix_origem"], col_offset=11
    )

    print("Lendo Mix de Origem diário (para filtro por período)...", file=sys.stderr)
    try:
        mix_origem_diario = build_mix_origem_diario(service, cfg["mix_obz_visitacao_id"])
    except Exception as e:
        # NUNCA deixa esse filtro novo derrubar o resto do pipeline -- se der erro (aba
        # renomeada, layout mudou etc.), o filtro por período fica sem dado neste ciclo,
        # mas o resto do Mix de Origem (Mensal/Acumulado) continua normal.
        print(f"AVISO: falha ao ler Mix de Origem diário ({e}) -- filtro por período fica vazio neste ciclo.", file=sys.stderr)
        mix_origem_diario = {}

    print("Lendo Eventos...", file=sys.stderr)
    eventos = build_eventos(
        service, cfg["visitacao_parques_id"], cfg["sheet_names"]["eventos"]
    )

    print("Lendo Apoio Report...", file=sys.stderr)
    try:
        apoio_report = build_apoio_report(
            service, cfg["visitacao_parques_id"], cfg["sheet_names"].get("apoio_report", "APOIO REPORT")
        )
    except Exception as e:
        # Igual às outras abas defensivas: se a aba mudar de layout ou nome, a aba Report
        # do painel fica vazia neste ciclo em vez de derrubar o resto do pipeline.
        print(f"AVISO: falha ao ler Apoio Report ({e}) -- aba Report fica vazia neste ciclo.", file=sys.stderr)
        apoio_report = {"realizadoOntem": {}, "resumoMensal": {}}

    print("Lendo Histórico...", file=sys.stderr)
    historico_id = cfg.get("visitacao_historico_id", "")
    if historico_id and not historico_id.startswith("COLE_AQUI"):
        try:
            historico = build_historico(service, historico_id)
        except Exception as e:
            # NUNCA deixa a aba Histórico derrubar o resto do pipeline -- se der erro (id
            # errado, aba renomeada, planilha ainda não compartilhada etc.), essa aba fica
            # vazia neste ciclo e o resto do data.json sai normal.
            print(f"AVISO: falha ao ler Histórico ({e}) -- aba fica vazia neste ciclo.", file=sys.stderr)
            historico = {"parques": {}, "parquesComShare": []}
    else:
        print(
            "Histórico: visitacao_historico_id ainda não configurado (placeholder) "
            "-- aba fica vazia até o ID real ser colado no config.json.",
            file=sys.stderr,
        )
        historico = {"parques": {}, "parquesComShare": []}

    print("Lendo Plano de Mídia...", file=sys.stderr)
    plano_midia_id = cfg.get("plano_midia_id", "")
    if plano_midia_id and not plano_midia_id.startswith("COLE_AQUI"):
        try:
            plano_midia = build_plano_midia(service, plano_midia_id)
        except Exception as e:
            print(f"AVISO: falha ao ler Plano de Mídia ({e}) -- aba fica vazia neste ciclo.", file=sys.stderr)
            plano_midia = {"resumo": {}, "detail": {}}
    else:
        print(
            "Plano de Mídia: plano_midia_id ainda não configurado (placeholder) "
            "-- aba fica vazia até o ID real ser colado no config.json.",
            file=sys.stderr,
        )
        plano_midia = {"resumo": {}, "detail": {}}

    output = {
        "geradoEm": datetime.datetime.utcnow().isoformat() + "Z",
        "VISITACAO": visitacao,
        "CAPTACAO_CV_3P_BY_MONTH": cv3p_by_month,
        "CAPTACAO_CV_3P_ANUAL": cv3p_anual,
        "CAPTACAO_CV_3P_DIARIO": cv3p_diario,
        "CAPTACAO_PNI_SEMMORADOR": semmorador_ratio,
        "SHARE_META_MESES": [MESES_PT[MONTH_NUMBER[m] - 1] for m in cfg["meses_com_dados"]],
        "SHARE_META_GRUPO_CATARATAS": share_meta_grupo_cataratas,
        "SHARE": {
            "dashShareGC": dash_share_gc,
            "investimentoMidia": {"meses": investimento_midia},
            "evolucaoMensal": evolucao_mensal,
        },
        "INVEST_MKT_RESUMO": invest_mkt_resumo,
        "INVEST_MKT_DETAIL": invest_mkt_detail,
        "MIX_ORIGEM": mix_origem,
        "MIX_ORIGEM_ACUMULADO": mix_origem_acumulado,
        "MIX_ORIGEM_DIARIO": mix_origem_diario,
        "EVENTOS": eventos,
        "APOIO_REPORT": apoio_report,
        "HISTORICO": historico,
        "PLANO_MIDIA": plano_midia,
    }

    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"OK: {args.out} gerado.", file=sys.stderr)


if __name__ == "__main__":
    main()
