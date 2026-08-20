# -*- coding: utf-8 -*-
"""
Extensão do extract_data.py: Cross AquaFoz (vendas cruzadas AquaFoz / PNI / Marco das 3
Fronteiras)
============================================================================================

CONTEXTO
--------
Pedido do usuário (20/08/2026): trazer pro Performance Dashboard, numa aba nova ("Cross
AquaFoz"), o relatório que hoje ele monta manualmente (a partir da planilha "Vendas Cross
para AquaFoz 2026") pra enviar pra Camila -- duas visões:

  - Diária: divisão por canal de venda (Marco das 3 Fronteiras: Bilheteria/Trade balcão/
    Cross/Combo; Aeroporto; PNI: Cross/Totem/Combo), Total cross, Share de visitação AQF/
    PNI, Visitação AquaFoz/PNI, Captação AQF x PNI -- tudo do MÊS VIGENTE, dia a dia.
  - Mensal: série mês a mês (Jan-Dez) com Vendas total, Vendas Totem PNI, Vendas Combo PNI,
    Visitação AQF, Share AQF, Visitação PNI, Share PNI, Captação AQF x PNI.

FONTE: planilha Google Sheets "Vendas Cross para AquaFoz 2026"
(id em cfg["cross_aquafoz_id"], ver COMO INTEGRAR no fim do arquivo).

REVISÃO DE 20/08/2026 -- ANÁLISE PROFUNDA (troca de fonte, ver abaixo o porquê)
--------------------------------------------------------------------------------------------
A primeira versão deste módulo lia a aba "Venda mes a mes" (pra Mensal) e a aba com o NOME
DO MÊS VIGENTE, ex. "Agosto" (pra Diária). Essas duas abas são preenchidas/coladas à mão e
tinham um problema real: nenhuma das duas tinha uma coluna separada pra "Combo MF3" (combo
vendido no próprio Marco das 3 Fronteiras) -- ela aparecia misturada ou simplesmente ausente,
enquanto o dashboard real do usuário (Looker Studio "Vendas Cross AQF", print enviado por ele
em 20/08) mostra "Combo MF3" como categoria própria, com valores reais desde Janeiro/2026.

Comparando os prints do dashboard real com a planilha, a aba "Geral data" (239 linhas, uma
por DIA, do ano inteiro, colunas: Mês | Data | Bilheteria MF3 | Trade MF3 | Aeroporto |
Cross MF3 | Cross PNI | Totem PNI | Combo PNI | Combo MF3 | Total cross | Visitação PNI |
Capt. Diária PNI | Visitação AQF | Capt Total AQF x PNI) bate EXATAMENTE, dia a dia, com os
números que aparecem no gráfico "Venda diária por local" do dashboard real (confirmado, ex.:
08/01 Combo MF3 = 65, 12/01 Combo MF3 = 23 -- os mesmos dois valores que aparecem no print).
Ou seja: "Geral data" é a fonte viva que alimenta o dashboard de verdade -- por isso este
módulo foi reescrito pra usar SÓ ela, tanto pra Diária quanto pra Mensal, e não depende mais
de nenhuma aba com nome de mês (nem precisa mais que alguém crie a aba do mês seguinte à mão
-- "Geral data" já vem com o ano inteiro).

IMPORTANTE -- achado da análise, pra registro: comparando "Geral data" com as abas manuais
"Venda mes a mes" / "Anual Data" (que tinham valores levemente diferentes, ex. Agosto: Total
4.821 nessas abas manuais vs 4.806 somando "Geral data" no mesmo instante, com Totem PNI
1.586 vs 1.486, Combo 2.745 vs 2.649+194 Combo MF3 separado), essas duas abas manuais
DIVERGEM da "Geral data" -- principalmente nos meses mais antigos (Janeiro a Junho, onde a
diferença chega a ser de 2x no total, ex. Janeiro: 116 nas abas manuais vs 239 em "Geral
data"). Isso indica que "Venda mes a mes"/"Anual Data" foram preenchidas manualmente num
processo antigo, sem refletir a Combo MF3 nem os ajustes que "Geral data" já tem -- não são
fórmula viva. Recomendação: se o usuário quiser bater os dois relatórios, o ideal é ele
revisar/migrar aquelas duas abas manuais também a partir de "Geral data" (fonte única). Este
painel já usa só "Geral data" a partir de agora.

  - Esquema de "Geral data" é FIXO desde Janeiro/2026 (15 colunas incluindo Combo MF3 -- ao
    contrário das antigas abas por mês, cujo esquema mudava mês a mês) -- por isso as colunas
    ainda são lidas pelo NOME do cabeçalho (DIARIO_COLUNAS), não por posição fixa, como
    proteção extra caso alguém insira/reordene uma coluna no futuro.
  - A coluna "Mês" (texto, ex. "Jan", "Fev", "Agosto" -- abreviação inconsistente) NÃO é usada
    pra decidir o mês de cada linha -- usa-se o mês da própria coluna "Data" (célula de data
    real, não texto, ao contrário de outras planilhas do painel onde a data vem como texto em
    formato ambíguo -- aqui já vem como datetime nativo do Google Sheets/Excel).
  - Um dia só entra nas somas (Mensal e Diária) se "Visitação PNI" daquele dia não estiver
    vazio -- dias futuros (do mês vigente ou de meses seguintes) vêm com tudo em branco na
    planilha (a própria "Total cross" da planilha soma como 0, "Capt. Diária PNI" vem como
    erro "#DIV/0!") e não devem contar como "dia com dado".
"""

import datetime
import unicodedata

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


def _norm(s):
    """Normaliza rótulo de célula (cabeçalho de coluna) pra comparação: remove acento, baixa
    a caixa, colapsa espaço. Não usado para os VALORES numéricos -- ver _to_num."""
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return " ".join(s.lower().split())


def _to_num(v):
    """Converte valor de célula pra número, tratando 'NA'/'N/A'/'#DIV/0!'/vazio como None (em
    vez de virar 0 escondido ou propagar erro) -- diferença importante pro front-end saber
    quando o dado simplesmente não existe ainda (ex.: dia futuro sem visitação lançada)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s or s.upper() in ("NA", "N/A", "#DIV/0!", "#N/A", "#REF!", "#VALUE!"):
        return None
    try:
        return float(s)
    except ValueError:
        try:
            return float(s.replace(".", "").replace(",", "."))
        except ValueError:
            return None


# Nome da coluna (cabeçalho, linha 1) na aba "Geral data" -> chave no dict de saída por dia.
DIARIO_COLUNAS = {
    "bilheteria mf3": "bilheteriaMf3",
    "trade mf3": "tradeMf3",
    "aeroporto": "aeroporto",
    "cross mf3": "crossMf3",
    "cross pni": "crossPni",
    "totem pni": "totemPni",
    "combo pni": "comboPni",
    "combo mf3": "comboMf3",
    "total cross": "totalCross",
    "visitacao pni": "visitacaoPni",
    "capt. diaria pni": "captDiariaPni",
    "visitacao aqf": "visitacaoAqf",
    "capt total aqf x pni": "captTotalAqfPni",
}

# Campos somados (não os de percentual/razão, que são recalculados a partir das somas -- ver
# _aggregate_mes) ao agregar os dias de um mês.
_CAMPOS_SOMA = [
    "bilheteriaMf3", "tradeMf3", "aeroporto", "crossMf3", "crossPni",
    "totemPni", "comboPni", "comboMf3", "totalCross", "visitacaoPni", "visitacaoAqf",
]


def _ler_dias(rows):
    """Lê a aba 'Geral data' (lista de listas, uma por linha da planilha) e devolve uma lista
    de dicts, um por dia com dado real (Visitação PNI não vazio), ordenada por data, com uma
    chave 'data' (datetime.date) + as chaves de DIARIO_COLUNAS."""
    if not rows:
        return []
    header_row = rows[0]
    col_data = None
    col_map = {}
    for i, v in enumerate(header_row):
        nv = _norm(v)
        if nv == "data":
            col_data = i
        key = DIARIO_COLUNAS.get(nv)
        if key:
            col_map[i] = key
    if col_data is None or not col_map:
        raise ValueError("colunas esperadas (Data + métricas) não encontradas na aba 'Geral data'")

    dias = []
    for row in rows[1:]:
        if not row or col_data >= len(row):
            continue
        dt_raw = row[col_data]
        if isinstance(dt_raw, datetime.datetime):
            dt = dt_raw.date()
        elif isinstance(dt_raw, datetime.date):
            dt = dt_raw
        else:
            continue  # célula vazia/texto -- fora do range de dias reais
        entry = {"data": dt}
        for i, key in col_map.items():
            entry[key] = _to_num(row[i]) if i < len(row) else None
        if entry.get("visitacaoPni") is None:
            continue  # dia sem dado lançado ainda (mês vigente em andamento ou mês futuro)
        dias.append(entry)
    dias.sort(key=lambda e: e["data"])
    return dias


def _aggregate_mes(dias_do_mes):
    """Soma os dias de um mês e recalcula os percentuais/razões A PARTIR DAS SOMAS do mês
    (não a média dos percentuais diários) -- mesma fórmula usada dia a dia na própria
    planilha, confirmada célula a célula:
      shareVisitacaoAqf ('Share AQF' no dashboard real) = vendasTotal / visitacaoAqf
      captacaoPni        ('Share PNI' no dashboard real) = vendasTotal / visitacaoPni
      captacaoAqfPni     ('Captação AQF x PNI')          = visitacaoAqf / visitacaoPni
    """
    soma = {k: 0.0 for k in _CAMPOS_SOMA}
    for d in dias_do_mes:
        for k in _CAMPOS_SOMA:
            soma[k] += d.get(k) or 0

    vendas_total = soma["totalCross"]
    visitacao_aqf = soma["visitacaoAqf"]
    visitacao_pni = soma["visitacaoPni"]
    return {
        "vendasTotal": vendas_total,
        "vendasTotemPni": soma["totemPni"],
        "vendasComboPni": soma["comboPni"],
        "vendasComboMf3": soma["comboMf3"],
        "visitacaoAqf": visitacao_aqf,
        "shareVisitacaoAqf": (vendas_total / visitacao_aqf) if visitacao_aqf else None,
        "visitacaoPni": visitacao_pni,
        "captacaoPni": (vendas_total / visitacao_pni) if visitacao_pni else None,
        "captacaoAqfPni": (visitacao_aqf / visitacao_pni) if visitacao_pni else None,
    }


def build_cross_aquafoz_mensal(dias):
    """Agrupa os dias (lista de dicts de _ler_dias) por mês (mês da própria data, não da
    coluna-texto 'Mês') e devolve {mes: {...}} -- ver _aggregate_mes pros campos."""
    por_mes = {}
    for d in dias:
        mes = MESES_PT[d["data"].month - 1]
        por_mes.setdefault(mes, []).append(d)
    return {mes: _aggregate_mes(lista) for mes, lista in por_mes.items()}


def build_cross_aquafoz_diaria(dias):
    """Devolve {'YYYY-MM-DD': {...}} com os campos crus de TODOS os dias com dado (ano
    inteiro, não só o mês vigente) -- inclui comboMf3.

    REVISÃO 20/08/2026: antes esta função só devolvia o mês vigente (o front-end só mostrava
    "o mês atual"). O usuário pediu um seletor de período (mês específico, ano inteiro,
    personalizado) igual ao que existe no dashboard real dele e na aba Mix de Origem deste
    painel -- pra isso o front-end precisa ter TODOS os dias disponíveis em mãos, não só o
    mês vigente, e filtra pelo período escolhido na hora de renderizar (ver
    getCrossAquafozRange() no index.html)."""
    result = {}
    for d in dias:
        entry = {k: d.get(k) for k in DIARIO_COLUNAS.values()}
        result[d["data"].strftime("%Y-%m-%d")] = entry
    return result


def _download_workbook(drive_service, spreadsheet_id):
    """Baixa a planilha (Drive API) e abre com openpyxl -- mesmo padrão de
    top5_origem_extension.py (self-contained, não depende do cache/import de
    extract_data.py, pra evitar import circular)."""
    import io
    import openpyxl
    from googleapiclient.http import MediaIoBaseDownload

    meta = drive_service.files().get(
        fileId=spreadsheet_id, fields="mimeType", supportsAllDrives=True
    ).execute()
    is_native_sheet = meta["mimeType"] == "application/vnd.google-apps.spreadsheet"

    buf = io.BytesIO()
    if is_native_sheet:
        request = drive_service.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive_service.files().get_media(fileId=spreadsheet_id, supportsAllDrives=True)
    downloader = MediaIoBaseDownload(buf, request, chunksize=10 * 1024 * 1024)
    done = False
    while not done:
        _, done = downloader.next_chunk(num_retries=5)
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True, read_only=True)


def _get_rows(wb, sheet_name):
    ws = wb[sheet_name]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def build_cross_aquafoz(drive_service, spreadsheet_id, mes_atual_nome):
    """Função principal -- chamar de dentro de main() do extract_data.py e salvar o resultado
    em data['CROSS_AQUAFOZ'].

    mes_atual_nome: nome do mês vigente em PT, Title Case (ex. 'Agosto') -- mesmo valor que já
    está em cfg["meses_com_dados"][-1] (MESES_PT) -- só ecoado no resultado (o front-end usa
    como valor padrão do seletor de período); a leitura em si é sempre da aba única
    'Geral data', ano inteiro.

    Devolve {'mensal': {...}, 'diaria': {...}, 'mesAtual': mes_atual_nome} -- 'diaria' traz
    TODOS os dias com dado do ano (não só o mês vigente, ver build_cross_aquafoz_diaria).
    """
    wb = _download_workbook(drive_service, spreadsheet_id)
    dias = _ler_dias(_get_rows(wb, "Geral data"))
    mensal = build_cross_aquafoz_mensal(dias)
    diaria = build_cross_aquafoz_diaria(dias)
    return {"mensal": mensal, "diaria": diaria, "mesAtual": mes_atual_nome}


# =========================================================================================
# COMO INTEGRAR no extract_data.py real
# =========================================================================================
# Nenhuma mudança de integração é necessária além da que já foi feita antes -- a assinatura
# de build_cross_aquafoz() não mudou (mesmos 3 argumentos, mesmo lugar de chamada em main()).
# Só troque o arquivo scripts/cross_aquafoz_extension.py por este (mesmo nome, mesmo lugar).
#
# Resumo de como já está plugado (pra referência, caso precise recriar do zero):
# 1. scripts/cross_aquafoz_extension.py (este arquivo) + import no topo do extract_data.py:
#      from cross_aquafoz_extension import build_cross_aquafoz
# 2. config.json (mesma pasta) já tem: "cross_aquafoz_id": "10y-MjciXhxYU7yBoUgymLmphuELd3SbyR6HbRejCGOo"
# 3. main() já chama build_cross_aquafoz(service, cross_aquafoz_id, mes_atual_nome) e salva em
#    output["CROSS_AQUAFOZ"].
# 4. A planilha "Vendas Cross para AquaFoz 2026" já está compartilhada com a service account
#    do pipeline (senão a leitura cai no except e a aba fica vazia -- ver aviso no log da
#    Action).
# =========================================================================================
